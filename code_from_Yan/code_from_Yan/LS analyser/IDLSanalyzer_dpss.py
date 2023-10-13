import numpy as np
import sys
import os
import time as tm
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QListWidgetItem, QLineEdit, QLabel, QRadioButton, QGridLayout, QPushButton, QAction, QActionGroup, QMenu, QInputDialog, qApp, QVBoxLayout
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from uiLSanalysis import Ui_IDLS_analyzer_dpss
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
from semiconductor.electrical.ionisation import Ionisation as Ion
from semiconductor.material.intrinsic_carrier_density import IntrinsicCarrierDensity as NI
from semiconductor.material.thermal_velocity import ThermalVelocity as the_vel
from semiconductor.general_functions import carrierfunctions as CF
from semiconductor.recombination.intrinsic import Radiative as rad
from semiconductor.recombination.intrinsic import Auger as aug
from scipy.optimize import curve_fit, minimize
import scipy.constants as const
import Newton as nt
import pandas as pd
import openpyxl as xls


class rawdata(object):
    def __init__(self, name=None, Ndop=None, temp=None, doptype=None, nxc=None, tau=None, PCPL='PC', thickness=None, j0=None, plotopt=True, uid=None, **kwarg):
        self.name = name
        self.Ndop = Ndop
        self.temp = temp
        self.doptype = doptype
        self.nxc = nxc
        self.tau = tau
        self.uid = uid
        self.plotopt = plotopt
        self.PCPL = PCPL
        self.thickness = thickness
        self.j0 = j0


class fitres(object):
    def __init__(self, name=None, Ndop=None, temp=None, doptype=None, m=None, b=None, parentid=None, brotherid=None, uid=None, **kwarg):
        self.name = name
        self.Ndop = Ndop
        self.temp = temp
        self.doptype = doptype
        self.m = m
        self.b = b
        self.uid = uid
        self.parentid = parentid
        self.brotherid = brotherid


class LSana_dpss(QMainWindow, Ui_IDLS_analyzer_dpss):

    def __init__(self, rawdata, uidlist, mainwindow, parent=None, **kwarg):

        QMainWindow.__init__(self, parent)
        self.setupUi(self)

        self.Rawdat = rawdata
        self.uidlist = uidlist
        self.mainwindow = mainwindow
        self.currentfitres = None
        self.fitreslist = []
        self.currentbroid = 0

        self.listWidget_fit.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection)
        self.listWidget_fit.itemSelectionChanged.connect(self.enablewidgts)
        self.listWidget_fit.clicked.connect(
            self.listWidget_fitres.clearSelection)
        self.listWidget_fit.itemSelectionChanged.connect(self.updateplot)
        self.listWidget_fit.itemDoubleClicked.connect(self.opensetparam2)
        self.listWidget_fit.setContextMenuPolicy(Qt.CustomContextMenu)
        self.listWidget_fit.customContextMenuRequested.connect(
            self.itemrightclicked2)

        self.listWidget_fitres.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection)
        self.listWidget_fitres.itemSelectionChanged.connect(self.enablewidgts)
        self.listWidget_fitres.clicked.connect(
            self.listWidget_fit.clearSelection)
        self.listWidget_fitres.itemSelectionChanged.connect(self.updateplot)
        self.listWidget_fitres.setContextMenuPolicy(Qt.CustomContextMenu)
        self.listWidget_fitres.customContextMenuRequested.connect(
            self.itemrightclicked)

        self.comboBox_plotopt2.currentIndexChanged.connect(self.updateplot)

        self.lineEdit_fitini1.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_fitini2.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_Etlb.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_Etub.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_dEt.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_klb.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_kup.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_dk.textChanged[str].connect(self.enablewidgts)
        self.lineEdit_simufitk.textChanged[str].connect(self.enablewidgts)

        self.lineEdit_j0.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_j0l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_j0u.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1Et.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1Etl.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1Etu.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1k.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1kl.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1ku.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1tau.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1taul.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d1tauu.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2Et.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2Etl.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2Etu.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2k.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2kl.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2ku.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2tau.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2taul.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d2tauu.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3Et1.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3Et1l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3Et1u.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3Et2.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3Et2l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3Et2u.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3k1.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3k1l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3k1u.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3k2.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3k2l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3k2u.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3tau1.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3tau1l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3tau1u.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3tau2.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3tau2l.textChanged[str].connect(self.advancedcheck)
        self.lineEdit_d3tau2u.textChanged[str].connect(self.advancedcheck)

        self.checkBox_intrinsic.toggled.connect(self.advancedcheck)
        self.checkBox_j0.toggled.connect(self.advancedcheck)
        self.checkBox_def1.toggled.connect(self.advancedcheck)
        self.checkBox_def2.toggled.connect(self.advancedcheck)
        self.checkBox_def3.toggled.connect(self.advancedcheck)

        self.checkBox_fixj0.toggled.connect(self.advancedcheck)
        self.checkBox_fixd1Et.toggled.connect(self.advancedcheck)
        self.checkBox_fixd1k.toggled.connect(self.advancedcheck)
        self.checkBox_fixd2Et.toggled.connect(self.advancedcheck)
        self.checkBox_fixd2k.toggled.connect(self.advancedcheck)
        self.checkBox_fixd3Et1.toggled.connect(self.advancedcheck)
        self.checkBox_fixd3Et2.toggled.connect(self.advancedcheck)
        self.checkBox_fixd3k1.toggled.connect(self.advancedcheck)
        self.checkBox_fixd3k2.toggled.connect(self.advancedcheck)

        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.verticalLayout.addWidget(self.canvas)
        self.verticalLayout.addWidget(self.toolbar)
        self.ax1 = self.figure.add_subplot(111)
        self.ax1.set_xlabel(r'Excess carrier density $[cm^{-3}]$')
        self.ax1.set_ylabel('Lifetime [s]')
        self.figure.tight_layout()

        self.pushButton_back.clicked.connect(self.close)

        self.pushButton_load.clicked.connect(self.loaddata)

        self.pushButton_set2.clicked.connect(self.opensetparam2)
        self.pushButton_set2.setEnabled(False)

        self.pushButton_dpss.clicked.connect(self.dpss)
        self.pushButton_dpss.setEnabled(False)

        self.pushButton_twofit.clicked.connect(self.twofit)
        self.pushButton_twofit.setEnabled(False)

        self.pushButton_1dplot.clicked.connect(self.onedplot)
        self.pushButton_1dplot.setEnabled(False)

        self.pushButton_onefit.clicked.connect(self.onefit)
        self.pushButton_onefit.setEnabled(False)

        self.pushButton_2dplot.clicked.connect(self.twodplot)
        self.pushButton_2dplot.setEnabled(False)

        self.pushButton_exportsimu.clicked.connect(self.exportsimu)
        self.pushButton_exportsimu.setEnabled(False)

        self.pushButton_simufit.clicked.connect(self.simufit)
        self.pushButton_simufit.setEnabled(False)

        self.pushButton_acceptfit.clicked.connect(self.acceptfit)
        self.pushButton_acceptfit.setEnabled(False)

        self.pushButton_delfitres.clicked.connect(self.delfit)
        self.pushButton_delfitres.setEnabled(False)

        self.pushButton_showfitting.clicked.connect(self.showfitting)
        self.pushButton_showfitting.setEnabled(False)

        self.radioButton_ind.toggled.connect(self.enablewidgts)

        self.pushButton_expplot2.clicked.connect(self.expplot2)

        self.pushButton_advfit.clicked.connect(self.advfit)
        self.pushButton_advfit.setEnabled(False)

        self.pushButton_plotadvfit.clicked.connect(self.plotadvfit)
        self.pushButton_plotadvfit.setEnabled(False)

        self.pushButton_expadvfit.clicked.connect(self.expadvfit)
        self.pushButton_expadvfit.setEnabled(False)

        self.pushButton_loadadvfit.clicked.connect(self.loadadvfit)
        self.pushButton_loadadvfit.setEnabled(False)

        if self.Rawdat is not None:
            for data in self.Rawdat:
                if data.uid in self.uidlist:
                    dataitem = QListWidgetItem(parent=self.listWidget_fit)
                    dataitem.setText(data.name)
                    dataitem.setData(32, data.uid)
                    self.listWidget_fit.addItem(dataitem)
        else:
            self.Rawdat = []
            self.pushButton_back.setEnabled(False)

        availablNI = NI().available_models()
        availablIon = Ion().available_models()
        availablthervel = the_vel().available_models()
        availablRad = rad().available_models()
        availablAug = aug().available_models()

        self.menubar = self.menuBar()
        self.choosmodel = self.menubar.addMenu('Choose your models')
        self.nimodel = self.choosmodel.addMenu('ni models')
        self.ionmodel = self.choosmodel.addMenu('Ionisation')
        self.themodel = self.choosmodel.addMenu('thermal velocity')
        self.Radmodel = self.choosmodel.addMenu('Radiative models')
        self.Augmodel = self.choosmodel.addMenu('Auger models')

        self.nigroup = QActionGroup(self)
        self.iongroup = QActionGroup(self)
        self.thegroup = QActionGroup(self)
        self.radgroup = QActionGroup(self)
        self.auggroup = QActionGroup(self)

        for nimodel in availablNI:
            a = self.nigroup.addAction(QAction(nimodel, checkable=True))
            if nimodel == 'Couderc_2014':
                a.setChecked(True)
            self.nimodel.addAction(a)
        for ionmodel in availablIon:
            a = self.iongroup.addAction(QAction(ionmodel, checkable=True))
            if ionmodel == 'Altermatt_2006_table1':
                a.setChecked(True)
            self.ionmodel.addAction(a)
        for themodel in availablthervel:
            a = self.thegroup.addAction(QAction(themodel, checkable=True))
            if themodel == 'Green_1990':
                a.setChecked(True)
            self.themodel.addAction(a)
        for radmodel in availablRad:
            b = self.radgroup.addAction(QAction(radmodel, checkable=True))
            if radmodel == 'Altermatt_2005':
                b.setChecked(True)
            self.Radmodel.addAction(b)
        for augmodel in availablAug:
            c = self.auggroup.addAction(QAction(augmodel, checkable=True))
            if augmodel == 'Richter2012':
                c.setChecked(True)
            self.Augmodel.addAction(c)

    def closeEvent(self, event):
        if self.mainwindow:
            self.mainwindow.show()

    def updateplot(self):
        self.ax1.clear()
        self.ax1.grid()
        for item in self.listWidget_fit.selectedItems():
            for data in self.Rawdat:
                if data.uid == item.data(32):
                    if self.comboBox_plotopt2.currentIndex() == 0:
                        X = self.getXY(nxc=data.nxc, T=data.temp,
                                       Ndop=data.Ndop, doptype=data.doptype)
                        self.ax1.set_xlabel(r'X/Y')
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.plot(X, data.tau, '.', label=data.name)
                    elif self.comboBox_plotopt2.currentIndex() == 1:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, data.tau, '.', label=data.name)
                    elif self.comboBox_plotopt2.currentIndex() == 2:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, 1. / data.tau,
                                      '.', label=data.name)

        for item in self.listWidget_fitres.selectedItems():
            for fitres in self.fitreslist:
                if fitres.uid == item.data(32):
                    sigleid = fitres.uid
                    broid = fitres.brotherid
                    praid = fitres.parentid

            for data in self.Rawdat:
                if data.uid == praid:
                    X = self.getXY(nxc=data.nxc, T=data.temp,
                                   Ndop=data.Ndop, doptype=data.doptype)
                    if self.comboBox_plotopt2.currentIndex() == 0:
                        self.ax1.plot(X, data.tau, '.', label=data.name)
                        if broid == None:
                            for fitres2 in self.fitreslist:
                                if fitres2.uid == sigleid:
                                    self.ax1.plot(
                                        X, X * fitres2.m + fitres2.b, label=fitres2.name)
                        else:
                            twolist = []
                            for fitres3 in self.fitreslist:
                                if fitres3.brotherid == broid:
                                    twolist.append(fitres3)
                                    self.ax1.plot(
                                        X, X * fitres3.m + fitres3.b, label=fitres3.name)
                            if len(twolist) == 2:
                                self.ax1.plot(X, 1 / (1. / (X * twolist[0].m + twolist[0].b) + 1 / (
                                    X * twolist[1].m + twolist[1].b)), label='Two defects fit' + data.name)
                        self.ax1.set_xlabel(r'X/Y')
                        self.ax1.set_ylabel('Lifetime [s]')

                    elif self.comboBox_plotopt2.currentIndex() == 1:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, data.tau, '.', label=data.name)
                        if broid == None:
                            for fitres2 in self.fitreslist:
                                if fitres2.uid == sigleid:
                                    self.ax1.plot(
                                        data.nxc, X * fitres2.m + fitres2.b, label=fitres2.name)
                        else:
                            twolist = []
                            for fitres3 in self.fitreslist:
                                if fitres3.brotherid == broid:
                                    twolist.append(fitres3)
                                    self.ax1.plot(
                                        data.nxc, X * fitres3.m + fitres3.b, label=fitres3.name)
                            if len(twolist) == 2:
                                self.ax1.plot(data.nxc, 1 / (1. / (X * twolist[0].m + twolist[0].b) + 1 / (
                                    X * twolist[1].m + twolist[1].b)), label='Two defects fit' + data.name)
                    elif self.comboBox_plotopt2.currentIndex() == 2:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, 1. / data.tau,
                                      '.', label=data.name)
                        if broid == None:
                            for fitres2 in self.fitreslist:
                                if fitres2.uid == sigleid:
                                    self.ax1.plot(
                                        data.nxc, 1. / (X * fitres2.m + fitres2.b), label=fitres2.name)
                        else:
                            twolist = []
                            for fitres3 in self.fitreslist:
                                if fitres3.brotherid == broid:
                                    twolist.append(fitres3)
                                    self.ax1.plot(
                                        data.nxc, 1. / (X * fitres3.m + fitres3.b), label=fitres3.name)
                            if len(twolist) == 2:
                                self.ax1.plot(data.nxc, (1. / (X * twolist[0].m + twolist[0].b) + 1 / (
                                    X * twolist[1].m + twolist[1].b)), label='Two defects fit' + data.name)

        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()

    def enablewidgts(self):
        if len(self.listWidget_fit.selectedItems()) == 1:
            self.pushButton_set2.setEnabled(True)
        else:
            self.pushButton_set2.setEnabled(False)
        if len(self.listWidget_fit.selectedItems()) == 1 and self.radioButton_ind.isChecked() == True:
            self.pushButton_onefit.setEnabled(True)
            try:
                a = float(self.lineEdit_fitini1.text())
                b = float(self.lineEdit_fitini2.text())
                if a < 100 and a > 0 and b < 100 and b > 0:
                    self.pushButton_twofit.setEnabled(True)
                else:
                    self.pushButton_twofit.setEnabled(False)
            except ValueError:
                self.pushButton_twofit.setEnabled(False)
        else:
            self.pushButton_onefit.setEnabled(False)
            self.pushButton_twofit.setEnabled(False)
        if len(self.listWidget_fitres.selectedItems()) != 0 and self.radioButton_ind.isChecked() == True:
            self.pushButton_delfitres.setEnabled(True)
            self.pushButton_dpss.setEnabled(True)
        else:
            self.pushButton_delfitres.setEnabled(False)
            self.pushButton_dpss.setEnabled(False)
        if len(self.listWidget_fit.selectedItems()) > 1 and self.radioButton_simu.isChecked() == True:
            try:
                if float(self.lineEdit_Etlb.text()) < float(self.lineEdit_Etub.text()) and float(self.lineEdit_klb.text()) < float(self.lineEdit_kup.text()) and float(self.lineEdit_klb.text()) > 0 and int(self.lineEdit_dEt.text()) > 0 and int(self.lineEdit_dk.text()) > 0:
                    self.pushButton_simufit.setEnabled(True)
                else:
                    self.pushButton_simufit.setEnabled(False)
            except ValueError:
                self.pushButton_simufit.setEnabled(False)
        else:
            self.pushButton_simufit.setEnabled(False)

        try:
            float(self.lineEdit_simufitEt.text())
            self.pushButton_1dplot.setEnabled(True)
            self.pushButton_2dplot.setEnabled(True)
            self.pushButton_exportsimu.setEnabled(True)
            self.pushButton_showfitting.setEnabled(True)
        except ValueError:
            self.pushButton_1dplot.setEnabled(False)
            self.pushButton_2dplot.setEnabled(False)
            self.pushButton_exportsimu.setEnabled(False)
            self.pushButton_showfitting.setEnabled(False)
        self.advancedcheck()

    def opensetparam2(self):
        self.dialog = QtWidgets.QDialog()
        grid = QGridLayout(self.dialog)
        self.ntype = QRadioButton('n-type')
        self.ptype = QRadioButton('p-type')
        self.dop = QLineEdit()
        self.temp = QLineEdit()
        self.jj0 = QLineEdit()
        self.thick = QLineEdit()
        Ldop = QLabel('Ndop (cm-3)')
        Ltemp = QLabel('Temp (K)')
        Lj0 = QLabel('J0 (A/cm2)')
        Lthick = QLabel('Thickness (cm)')
        self.ok = QPushButton('OK')
        self.ok.setEnabled(False)
        grid.addWidget(self.ntype, 0, 0)
        grid.addWidget(self.ptype, 0, 1)
        grid.addWidget(Ldop, 1, 0)
        grid.addWidget(self.dop, 1, 1)
        grid.addWidget(Ltemp, 2, 0)
        grid.addWidget(self.temp, 2, 1)
        grid.addWidget(Lj0, 3, 0)
        grid.addWidget(self.jj0, 3, 1)
        grid.addWidget(Lthick, 4, 0)
        grid.addWidget(self.thick, 4, 1)
        grid.addWidget(self.ok, 5, 1)
        for data in self.Rawdat:
            if data.uid == self.listWidget_fit.selectedItems()[0].data(32):
                if data.Ndop is not None:
                    self.dop.setText('{:e}'.format(data.Ndop))
                if data.temp is not None:
                    self.temp.setText(str(data.temp))
                if data.doptype == 'n':
                    self.ntype.setChecked(True)
                if data.doptype == 'p':
                    self.ptype.setChecked(True)
                if data.j0 is not None:
                    self.jj0.setText('{:e}'.format(data.j0))
                if data.thickness is not None:
                    self.thick.setText(str(data.thickness))
        self.dop.textChanged[str].connect(self.checkparam2)
        self.temp.textChanged[str].connect(self.checkparam2)
        self.thick.textChanged[str].connect(self.checkparam2)
        self.jj0.textChanged[str].connect(self.checkparam2)
        self.ntype.toggled.connect(self.checkparam2)
        self.ok.clicked.connect(self.setparam2)
        self.dialog.exec_()

    def checkparam2(self):
        try:
            if float(self.dop.text()) > 0 and float(self.temp.text()) > 0 and len(self.listWidget_fit.selectedItems()) == 1:
                self.ok.setEnabled(True)
            else:
                self.ok.setEnabled(False)
        except ValueError:
            self.ok.setEnabled(False)
        if self.thick.text() != "":
            try:
                if float(self.thick.text()) < 0:
                    self.ok.setEnabled(False)
            except ValueError:
                self.ok.setEnabled(False)
        if self.jj0.text() != "":
            try:
                if float(self.jj0.text()) < 0:
                    self.ok.setEnabled(False)
            except ValueError:
                self.ok.setEnabled(False)

    def setparam2(self):
        for data in self.Rawdat:
            if data.uid == self.listWidget_fit.selectedItems()[0].data(32):
                data.Ndop = float(self.dop.text())
                data.temp = float(self.temp.text())
                if self.ntype.isChecked():
                    data.doptype = 'n'
                if self.ptype.isChecked():
                    data.doptype = 'p'
                try:
                    data.j0 = float(self.jj0.text())
                except ValueError:
                    data.j0 = None
                try:
                    data.thickness = float(self.thick.text())
                except ValueError:
                    data.thickness = None
        self.dialog.close()

    def itemrightclicked2(self):
        if len(self.listWidget_fit.selectedItems()) == 1:
            menu = QMenu(self.listWidget_fit)
            Rset = menu.addAction('Set')
            Rset.triggered.connect(self.opensetparam2)
            Rrename = menu.addAction('Rename')
            Rrename.triggered.connect(self.changename2)
            menu.popup(QtGui.QCursor.pos())

    def changename2(self):
        if len(self.listWidget_fit.selectedItems()) == 1:
            item = self.listWidget_fit.selectedItems()[0]
            text, ok = QInputDialog.getText(
                self, 'Rename', 'Enter a new name:')
            if ok:
                item.setText(text)
                for data in self.Rawdat:
                    if data.uid == item.data(32):
                        data.name = text
                self.updateplot()

    def getXY(self, nxc, T, Ndop, doptype, **kwarg):
        for action in self.ionmodel.actions():
            if action.isChecked():
                ionauthor = action.text()
        for action in self.nimodel.actions():
            if action.isChecked():
                ni_author = action.text()
        if doptype == 'n':
            Nidop = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=nxc, impurity='phosphorous', author=ionauthor)
            ne, nh = CF.get_carriers(
                0, Nidop, nxc, temp=T, ni_author=ni_author)
            X = np.divide(nxc, ne)
        elif doptype == 'p':
            Nidop = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=nxc, impurity='boron', author=ionauthor)
            ne, nh = CF.get_carriers(
                Nidop, 0, nxc, temp=T, ni_author=ni_author)
            X = np.divide(nxc, nh)
        return X

    def twoline(self, x, s1, t1, s2, t2):
        return np.divide(1, np.divide(1, s1 * x + t1) + np.divide(1, s2 * x + t2))

    def twolinenew(self, x, s1, t1, s2, t2):
        return np.divide(1, np.divide(1, (s1 - t1) * x + t1) + np.divide(1, (s2 - t2) * x + t2))

    def twofit(self):
        self.pushButton_acceptfit.setEnabled(True)
        self.updateplot()
        item = self.listWidget_fit.selectedItems()[0]
        for data in self.Rawdat:
            if data.uid == item.data(32):
                data2fit = data
        X = self.getXY(nxc=data2fit.nxc, T=data2fit.temp,
                       Ndop=data2fit.Ndop, doptype=data2fit.doptype)
        index1 = int(X.shape[0] * float(self.lineEdit_fitini1.text()) / 100)
        index2 = int(X.shape[0] * float(self.lineEdit_fitini2.text()) / 100)
        s1, t1 = np.polyfit(X[:index1], data2fit.tau[:index1], 1)
        s2, t2 = np.polyfit(X[index2:], data2fit.tau[index2:], 1)
        sigma = data2fit.tau/np.max(data2fit.tau)
        try:
            popt, pcov = curve_fit(
                self.twoline, X, data2fit.tau, p0=[s1, t1, s2, t2],sigma=sigma)
            if popt[3] + popt[2] > popt[1] + popt[0]:
                m = [popt[0], popt[2]]
                b = [popt[1], popt[3]]
            elif popt[3] + popt[2] <= popt[1] + popt[0]:
                m = [popt[2], popt[0]]
                b = [popt[3], popt[1]]
            fittau = self.twoline(X, *popt)

            if b[0] <= 0 or b[1] <= 0 or m[0] + b[0] <= 0 or m[1] + b[1] <= 0:
                t11 = np.amax([t1, 0])
                s11 = np.amax([s1 + t1, 0])
                # s22, t22 = np.polyfit(X[index2:], tau[index2:], 1)
                t22 = np.amax([t2, 0])
                s22 = np.amax([s2 + t2, 0])
                popt, pcov = curve_fit(self.twolinenew, X, data2fit.tau, p0=[s11, t11, s22, t22], bounds=(
                    0, [np.inf, np.inf, np.inf, np.inf]),sigma=sigma)
                if popt[2] > popt[0]:
                    m = [popt[0] - popt[1], popt[2] - popt[3]]
                    b = [popt[1], popt[3]]
                elif popt[2] <= popt[0]:
                    m = [popt[2] - popt[3], popt[0] - popt[1]]
                    b = [popt[3], popt[1]]
                fittau = self.twolinenew(X, *popt)

            domi = m[0] * X + b[0]
            secon = m[1] * X + b[1]
            residuals = data2fit.tau - fittau
            res = np.sum(residuals**2)
            ss_tot = np.sum((data2fit.tau - np.mean(data2fit.tau))**2)
            r_squared = 1 - (res / ss_tot)
            if self.comboBox_plotopt2.currentIndex() == 0:
                self.ax1.plot(X, domi, label='Dominant')
                self.ax1.plot(X, secon, label='Secondary')
                self.ax1.plot(X, fittau, label='Two defects fit')
            elif self.comboBox_plotopt2.currentIndex() == 1:
                self.ax1.plot(data2fit.nxc, domi, label='Dominant')
                self.ax1.plot(data2fit.nxc, secon, label='Secondary')
                self.ax1.plot(data2fit.nxc, fittau,
                              label='1 defect fit' + data2fit.name)
            elif self.comboBox_plotopt2.currentIndex() == 2:
                self.ax1.plot(data2fit.nxc, 1. / domi, label='Dominant')
                self.ax1.plot(data2fit.nxc, 1. / secon, label='Secondary')
                self.ax1.plot(data2fit.nxc, 1. / fittau,
                              label='1 defect fit' + data2fit.name)
            self.ax1.set_title(
                'Residual={:e}, R2={:f}'.format(res, r_squared), fontsize=10)
            self.ax1.legend(loc=0)
            self.figure.tight_layout()
            self.canvas.draw()
            aa = self.generatebroid()
            self.currentfitres = [fitres(
                name='1st defect_' + data2fit.name, Ndop=data2fit.Ndop, temp=data2fit.temp, doptype=data2fit.doptype, m=m[0], b=b[0], brotherid=aa, parentid=data2fit.uid), fitres(
                    name='2nd defect_' + data2fit.name, Ndop=data2fit.Ndop, temp=data2fit.temp, doptype=data2fit.doptype, m=m[1], b=b[1], brotherid=aa, parentid=data2fit.uid)]

        except RuntimeError:
            self.ax1.set_title(
                'Cannot fit, try to change the initial fit range or go back to data process', fontsize=10)

    def generatebroid(self):
        self.currentbroid += 1
        return (self.currentbroid - 1)

    def onefit(self):
        self.pushButton_acceptfit.setEnabled(True)
        self.updateplot()
        item = self.listWidget_fit.selectedItems()[0]
        for data in self.Rawdat:
            if data.uid == item.data(32):
                data2fit = data
        X = self.getXY(nxc=data2fit.nxc, T=data2fit.temp,
                       Ndop=data2fit.Ndop, doptype=data2fit.doptype)
        p = np.polyfit(X, data2fit.tau, 1)
        m = p[0]
        b = p[1]
        fittau = m * X + b
        residuals = fittau - data2fit.tau
        res = np.sum(residuals**2)
        ss_tot = np.sum((data2fit.tau - np.mean(data2fit.tau))**2)
        r_squared = 1 - (res / ss_tot)
        if self.comboBox_plotopt2.currentIndex() == 0:
            self.ax1.plot(
                X, fittau, label='1 defect fit' + data2fit.name)
        elif self.comboBox_plotopt2.currentIndex() == 1:
            self.ax1.plot(data2fit.nxc, fittau,
                          label='1 defect fit' + data2fit.name)
        elif self.comboBox_plotopt2.currentIndex() == 2:
            self.ax1.plot(data2fit.nxc, 1. / fittau,
                          label='1 defect fit' + data2fit.name)
        self.ax1.set_title(
            'Residual={:e}, R2={:f}'.format(res, r_squared), fontsize=10)
        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()
        self.currentfitres = [fitres(
            name='single defect_' + data2fit.name, Ndop=data2fit.Ndop, temp=data2fit.temp, doptype=data2fit.doptype, m=m, b=b, parentid=data2fit.uid)]

    def acceptfit(self):
        self.pushButton_acceptfit.setEnabled(False)
        self.fitreslist.extend(self.currentfitres)
        self.updateuid()
        self.updatefitreslist()

    def updatefitreslist(self):
        self.listWidget_fitres.clear()
        for data in self.fitreslist:
            dataitem = QListWidgetItem(parent=self.listWidget_fitres)
            dataitem.setText(data.name)
            dataitem.setData(32, data.uid)
            self.listWidget_fitres.addItem(dataitem)

    def updateuid(self):
        for fitres in self.fitreslist:
            fitres.uid = self.fitreslist.index(fitres)

    def SRHtau(self, nxc, Et, tau_e, tau_h, T, Na, Nd, ni_author, **kwargs):
        ni = NI().update(temp=T, author=ni_author)
        nh1 = ni * np.exp(-Et * const.e / (const.k * T))
        ne1 = ni * np.exp(Et * const.e / (const.k * T))
        ne, nh = CF.get_carriers(Na=Na, Nd=Nd, nxc=nxc, temp=T, ni=ni)
        U = (ne * nh - ni**2) / (tau_h * (ne + ne1) + tau_e * (nh + nh1))
        return nxc / U

    def SRHtau2(self, nxc, Et, tau_m, k, T, Ndop, doptype, ni_author, vth_author, ionauthor, **kwargs):
        vth_e300, vth_h300 = the_vel().update(temp=300, author=vth_author)
        ni = NI().update(temp=T, author=ni_author)
        nh1 = ni * np.exp(-Et * const.e / (const.k * T))
        ne1 = ni * np.exp(Et * const.e / (const.k * T))
        vth_e, vth_h = the_vel().update(temp=T, author=vth_author)
        if doptype == 'n':
            Nd = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=nxc, impurity='phosphorous', author=ionauthor)
            Na = 0
            tau_e = tau_m * vth_h300 / vth_e / k
            tau_h = tau_m * vth_h300 / vth_h
        elif doptype == 'p':
            Na = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=nxc, impurity='boron', author=ionauthor)
            Nd = 0
            tau_e = tau_m * vth_e300 / vth_e
            tau_h = tau_m * k * vth_e300 / vth_h

        ne, nh = CF.get_carriers(Na=Na, Nd=Nd, nxc=nxc, temp=T, ni=ni)
        U = (ne * nh - ni**2) / (tau_h * (ne + ne1) + tau_e * (nh + nh1))
        return nxc / U

    def SRHtau3(self, nxc, d3Et1, d3Et2, d3tau1, d3tau2, d3k1, d3k2, T, Ndop, doptype, ni_author, vth_author, ionauthor, **kwargs):
        vth_e300, vth_h300 = the_vel().update(temp=300, author=vth_author)
        ni = NI().update(temp=T, author=ni_author)
        nh1 = ni * np.exp(-d3Et1 * const.e / (const.k * T))
        ne1 = ni * np.exp(d3Et1 * const.e / (const.k * T))
        nh2 = ni * np.exp(-d3Et2 * const.e / (const.k * T))
        ne2 = ni * np.exp(d3Et2 * const.e / (const.k * T))
        vth_e, vth_h = the_vel().update(temp=T, author=vth_author)
        if doptype == 'n':
            Nd = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=nxc, impurity='phosphorous', author=ionauthor)
            Na = 0
            tau_e1 = d3tau1 * vth_h300 / vth_e / d3k1
            tau_h1 = d3tau1 * vth_h300 / vth_h
            tau_e2 = d3tau2 * vth_h300 / vth_e / d3k2
            tau_h2 = d3tau2 * vth_h300 / vth_h
        elif doptype == 'p':
            Na = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=nxc, impurity='boron', author=ionauthor)
            Nd = 0
            tau_e1 = d3tau1 * vth_e300 / vth_e
            tau_h1 = d3tau1 * d3k1 * vth_e300 / vth_h
            tau_e2 = d3tau2 * vth_e300 / vth_e
            tau_h2 = d3tau2 * d3k2 * vth_e300 / vth_h
        ne, nh = CF.get_carriers(Na=Na, Nd=Nd, nxc=nxc, temp=T, ni=ni)
        U = (1 / tau_e1 * ((ne * nh / tau_h1 - ne1 * nh1 / tau_h1) / (ne / tau_e1 + nh1 / tau_h1)) + 1 / tau_e2 * ((ne * nh / tau_h2 - ne2 * nh2 / tau_h2) / (ne2 /
                                                                                                                                                              tau_e2 + nh / tau_h2))) / (1 + ((ne1 / tau_e1 + nh / tau_h1) / (ne / tau_e1 + nh1 / tau_h1)) + ((ne / tau_e2 + nh2 / tau_h2) / (ne2 / tau_e2 + nh / tau_h2)))
        return nxc / U

    def minfunction_sameTaum(self, taum, nxclist, taulist, Et, k, Ndoplist, doptypelist, Tlist, ni_author, vth_author, ionauthor, ** kwarg):
        res = 0
        vth_e300, vth_h300 = the_vel().update(temp=300, author=vth_author)
        for nxc, tau, Ndop, doptype, T in zip(nxclist, taulist, Ndoplist, doptypelist, Tlist):
            vth_e, vth_h = the_vel().update(temp=T, author=vth_author)
            if doptype == 'n':
                Nd = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                    N_dop=Ndop, nxc=nxc, impurity='phosphorous', author=ionauthor)
                Na = 0
                tausimu = self.SRHtau(nxc=nxc, Et=Et, tau_e=taum * vth_h300 / vth_e / k,
                                      tau_h=taum * vth_h300 / vth_h, T=T, Na=Na, Nd=Nd, ni_author=ni_author)
            elif doptype == 'p':
                Na = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                    N_dop=Ndop, nxc=nxc, impurity='boron', author=ionauthor)
                Nd = 0
                tausimu = self.SRHtau(nxc=nxc, Et=Et, tau_e=taum * vth_e300 / vth_e,
                                      tau_h=taum * k * vth_e300 / vth_h, T=T, Na=Na, Nd=Nd, ni_author=ni_author)
            res += np.sum(np.abs((np.asarray(tau) - np.asarray(tausimu)
                                  ) / np.asarray(tau))) / len(nxc)
        res /= len(nxclist)
        return res

    def simufit_variTaum(self, Etlist, klist, nxclist, taulist, Ndoplist, doptypelist, Tlist, ni_author, vth_author, ionauthor,  **kwarg):
        ctperc = 0
        x0 = [1e-6]
        bounds = [(0, 1)]
        taummap = np.zeros((len(Etlist), len(klist), len(nxclist)))
        residualmap = np.zeros((len(Etlist), len(klist)))
        for m in range(len(Etlist)):
            Et = Etlist[m]
            for n in range(len(klist)):
                k = klist[n]
                for t in range(len(nxclist)):
                    nxc = nxclist[t]
                    tau = taulist[t]
                    Ndop = Ndoplist[t]
                    doptype = doptypelist[t]
                    T = Tlist[t]
                    fitres = minimize(self.minfunction_sameTaum, x0=x0, args=(
                        [nxc], [tau], Et, k, [Ndop], [doptype], [T], ni_author, vth_author, ionauthor), bounds=bounds)
                    taummap[m, n, t] = fitres.x
                    residualmap[m, n] += fitres.fun
                    ctperc += 1 / len(Etlist) / len(klist) / len(nxclist)
                    QApplication.processEvents()
                    self.progressBar.setValue(ctperc * 100)
        return taummap, residualmap

    def simufit_sameTaum(self, Etlist, klist, nxclist, taulist, Ndoplist, doptypelist, Tlist, ni_author, vth_author, ionauthor, **kwarg):
        ctperc = 0
        x0 = [1e-6]
        bounds = [(0, 1)]
        taummap = np.zeros((len(Etlist), len(klist)))
        residualmap = np.zeros((len(Etlist), len(klist)))
        for m in range(len(Etlist)):
            Et = Etlist[m]
            for n in range(len(klist)):
                k = klist[n]
                fitres = minimize(self.minfunction_sameTaum, x0=x0, args=(
                    nxclist, taulist, Et, k, Ndoplist, doptypelist, Tlist, ni_author, vth_author, ionauthor), bounds=bounds)
                taummap[m, n] = fitres.x
                residualmap[m, n] = fitres.fun
                ctperc += 1 / len(Etlist) / len(klist)
                QApplication.processEvents()
                self.progressBar.setValue(ctperc * 100)
        return taummap, residualmap

    def simufit(self):
        for item in self.listWidget_fit.selectedItems():
            for data in self.Rawdat:
                if data.uid == item.data(32):
                    dataitem = QListWidgetItem(
                        parent=self.listWidget_simufitlist)
                    dataitem.setText(data.name)
                    dataitem.setData(32, data.uid)
                    self.listWidget_simufitlist.addItem(dataitem)
        self.currentsimufituid = []
        self.progressBar.setValue(0)
        for action in self.ionmodel.actions():
            if action.isChecked():
                self.simuionauthor = action.text()
        for action in self.nimodel.actions():
            if action.isChecked():
                self.simuni_author = action.text()
        for action in self.themodel.actions():
            if action.isChecked():
                self.simuvth_author = action.text()
        self.Etscanlist = np.linspace(float(self.lineEdit_Etlb.text()), float(
            self.lineEdit_Etub.text()), int(self.lineEdit_dEt.text()))
        self.kscanlist = np.linspace(float(self.lineEdit_klb.text()), float(
            self.lineEdit_kup.text()), int(self.lineEdit_dk.text()))
        Ndoplist = []
        Tlist = []
        doptypelist = []
        nxclist = []
        taulist = []
        for item in self.listWidget_fit.selectedItems():
            for data in self.Rawdat:
                if data.uid == item.data(32):
                    self.currentsimufituid.append(data.uid)
                    Ndoplist.append(data.Ndop)
                    Tlist.append(data.temp)
                    doptypelist.append(data.doptype)
                    nxclist.append(data.nxc)
                    taulist.append(data.tau)
        if self.checkBox_sametaum.isChecked():
            self.taummap, self.residualmap = self.simufit_sameTaum(Etlist=self.Etscanlist, klist=self.kscanlist, nxclist=nxclist, taulist=taulist,
                                                                   Ndoplist=Ndoplist, doptypelist=doptypelist, Tlist=Tlist, ni_author=self.simuni_author, vth_author=self.simuvth_author, ionauthor=self.simuionauthor)
        else:
            self.taummap, self.residualmap = self.simufit_variTaum(Etlist=self.Etscanlist, klist=self.kscanlist, nxclist=nxclist, taulist=taulist,
                                                                   Ndoplist=Ndoplist, doptypelist=doptypelist, Tlist=Tlist, ni_author=self.simuni_author, vth_author=self.simuvth_author, ionauthor=self.simuionauthor)
        self.index = np.unravel_index(
            np.argmin(self.residualmap), self.residualmap.shape)
        self.lineEdit_simufitEt.setText(str(self.Etscanlist[self.index[0]]))
        self.lineEdit_simufitk.setText(str(self.kscanlist[self.index[1]]))
        self.lineEdit_simufittauminor.setText(str(self.taummap[self.index]))
        self.showfitting()

    def onedplot(self):
        self.dialog1dplot = QtWidgets.QDialog()
        self.onedvlayout = QVBoxLayout(self.dialog1dplot)
        self.figure3 = plt.figure()
        self.canvas3 = FigureCanvas(self.figure3)
        self.toolbar3 = NavigationToolbar(self.canvas3, self)
        self.onedvlayout.addWidget(self.canvas3)
        self.onedvlayout.addWidget(self.toolbar3)
        self.ax5 = self.figure3.add_subplot(211)
        self.ax6 = self.figure3.add_subplot(212)
        self.ax5.set_xlabel(r'$E_{t}-E_{i}$ [eV]')
        self.ax5.set_ylabel('Residual')
        self.ax6.set_xlabel('k')
        self.ax6.set_ylabel('Residual')
        self.ax5.semilogy()
        self.ax6.semilogy()
        self.ax5.plot(self.Etscanlist, self.residualmap[:, self.index[1]])
        self.ax6.plot(self.kscanlist, self.residualmap[self.index[0], :])
        self.figure3.tight_layout()
        self.dialog1dplot.exec_()

    def twodplot(self):
        self.dialog2dplot = QtWidgets.QDialog()
        self.twodvlayout = QVBoxLayout(self.dialog2dplot)
        self.figure4 = plt.figure()
        self.canvas4 = FigureCanvas(self.figure4)
        self.toolbar4 = NavigationToolbar(self.canvas4, self)
        self.twodvlayout.addWidget(self.canvas4)
        self.twodvlayout.addWidget(self.toolbar4)
        self.ax7 = self.figure4.add_subplot(111)
        self.ax7.set_xlabel(r'$E_{t}-E_{i}$ [eV]')
        self.ax7.set_ylabel('k')
        hh = self.ax7.imshow(self.residualmap.T, cmap='hot', aspect='auto', extent=[
            1.5 * self.Etscanlist[0] - 0.5 * self.Etscanlist[1], 1.5 * self.Etscanlist[-1] -
            0.5 * self.Etscanlist[-2], 1.5 * self.kscanlist[0] - 0.5 * self.kscanlist[1], 1.5 * self.kscanlist[-1] - 0.5 * self.kscanlist[-2]], origin='lower')
        # self.ax7.set_xticks(self.Etscanlist)
        # self.ax7.set_yticks(self.kscanlist)
        self.figure4.colorbar(hh)
        self.figure4.tight_layout()
        self.dialog2dplot.exec_()

    def showfitting(self):
        self.ax1.clear()
        self.ax1.grid()
        for simuid in self.currentsimufituid:
            for data in self.Rawdat:
                if data.uid == simuid:
                    if type(self.taummap[self.index]) is np.float64:
                        tau_m = self.taummap[self.index]
                    else:
                        tau_m = self.taummap[self.index][self.currentsimufituid.index(
                            simuid)]
                    tausimu = self.SRHtau2(nxc=data.nxc, Et=self.Etscanlist[self.index[0]], tau_m=tau_m, k=self.kscanlist[self.index[1]], T=data.temp, Ndop=data.Ndop, doptype=data.doptype, ni_author=self.simuni_author,
                                           vth_author=self.simuvth_author, ionauthor=self.simuionauthor)
                    if self.comboBox_plotopt2.currentIndex() == 0:
                        X = self.getXY(nxc=data.nxc, T=data.temp,
                                       Ndop=data.Ndop, doptype=data.doptype)
                        self.ax1.set_xlabel(r'X/Y')
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.plot(X, data.tau, '.', label=data.name)
                        self.ax1.plot(X, tausimu, label=data.name + 'fit')
                    elif self.comboBox_plotopt2.currentIndex() == 1:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, data.tau, '.', label=data.name)
                        self.ax1.plot(data.nxc, tausimu,
                                      label=data.name + 'fit')
                    elif self.comboBox_plotopt2.currentIndex() == 2:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, 1. / data.tau,
                                      '.', label=data.name)
                        self.ax1.plot(data.nxc, 1. / tausimu,
                                      label=data.name + 'fit')
        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()

    def itemrightclicked(self):
        if len(self.listWidget_fitres.selectedItems()) > 0:
            menu = QMenu(self.listWidget_fitres)
            Rdel = menu.addAction('Delete')
            Rdel.triggered.connect(self.delfit)
            if len(self.listWidget_fitres.selectedItems()) == 1:
                Rrename = menu.addAction('Rename')
                Rrename.triggered.connect(self.changename)
            menu.popup(QtGui.QCursor.pos())

    def delfit(self):
        for item in self.listWidget_fitres.selectedItems():
            for fitres in self.fitreslist:
                if fitres.uid == item.data(32):
                    self.fitreslist.remove(fitres)
        self.updateuid()
        self.updatefitreslist()
        self.updateplot()

    def changename(self):
        if len(self.listWidget_fitres.selectedItems()) == 1:
            item = self.listWidget_fitres.selectedItems()[0]
            text, ok = QInputDialog.getText(
                self, 'Rename', 'Enter a new name:')
            if ok:
                item.setText(text)
                for data in self.fitreslist:
                    if data.uid == item.data(32):
                        data.name = text
        self.updateplot()

    def dpss(self):
        self.dialogdpss = QtWidgets.QDialog()
        self.dialogdpss.setGeometry(400, 50, 600, 950)
        self.dpssvlayout = QVBoxLayout(self.dialogdpss)
        self.NTbuttom = QPushButton('Display Newton method result')
        self.NTbuttom.clicked.connect(self.NT)
        self.dpssvlayout.addWidget(self.NTbuttom)
        self.Exportbuttom = QPushButton('Export results')
        self.Exportbuttom.clicked.connect(self.exportind)
        self.dpssvlayout.addWidget(self.Exportbuttom)
        if len(self.listWidget_fitres.selectedItems()) < 2:
            self.NTbuttom.setEnabled(False)
        self.figure2 = plt.figure()
        self.canvas2 = FigureCanvas(self.figure2)
        self.toolbar2 = NavigationToolbar(self.canvas2, self)
        self.dpssvlayout.addWidget(self.canvas2)
        self.dpssvlayout.addWidget(self.toolbar2)
        self.ax3 = self.figure2.add_subplot(311)
        self.ax4 = self.figure2.add_subplot(312, sharex=self.ax3)
        self.ax44 = self.figure2.add_subplot(313, sharex=self.ax3)
        self.plotDPSS()
        self.dialogdpss.exec_()

    def plotDPSS(self):
        self.ax3.clear()
        self.ax4.clear()
        self.ax44.clear()
        self.ax44.set_xlabel(r'$E_{t}-E_{i}$ [eV]')
        self.ax4.set_ylabel('k')
        self.ax44.set_ylabel(r'Std Dev of the $E_{t}-k$ curves')
        self.ax3.set_ylabel(r'$\tau_{minor}$ [s]')
        self.ax3.semilogy()
        self.ax4.semilogy()
        self.ax44.semilogy()
        self.klist = []
        self.tauminorlist = []
        self.dpsshead1 = 'Et'
        self.dpsshead2 = ''
        for item in self.listWidget_fitres.selectedItems():
            for fitres in self.fitreslist:
                if fitres.uid == item.data(32):
                    EtRange, k, TauMinor = self.CalDPSS(
                        m=fitres.m, b=fitres.b, T=fitres.temp, Ndop=fitres.Ndop, doptype=fitres.doptype)
                    self.ax3.plot(EtRange, TauMinor, label=fitres.name)
                    self.ax4.plot(EtRange, k, label=fitres.name)
                    self.klist.append(k)
                    self.tauminorlist.append(TauMinor)
                    self.dpsshead1 += ',k' + fitres.name
                    self.dpsshead2 += ',TauMinor' + fitres.name
        self.Etlist = EtRange
        self.dpsshead2 += ',StdDev'
        klist = np.asarray(self.klist)
        self.stdk = np.std(klist, axis=0)/np.mean(klist, axis=0)
        self.ax44.plot(EtRange, self.stdk)
        self.ax4.legend(loc=0)
        self.figure2.subplots_adjust(left=0.13, bottom=0.06,
                                     right=0.9, top=0.96, wspace=0, hspace=0.23)

    def CalDPSS(self, m, b, T, Ndop, doptype, **kwarg):
        kb = const.k / const.e
        for action in self.ionmodel.actions():
            if action.isChecked():
                ionauthor = action.text()
        for action in self.nimodel.actions():
            if action.isChecked():
                ni_author = action.text()
        for action in self.themodel.actions():
            if action.isChecked():
                theauthor = action.text()
        vth_e300, vth_h300 = the_vel().update(temp=300, author=theauthor)
        EtRange = np.linspace(-0.6, 0.6, 401)
        k = np.ones(EtRange.shape[0])
        TauMinor = np.ones(EtRange.shape[0])
        vth_e, vth_h = the_vel().update(temp=T, author=theauthor)
        if doptype == 'n':
            Nidop = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=0, impurity='phosphorous', author=ionauthor)
            n0, p0 = CF.get_carriers(0, Nidop, 0, temp=T)
        elif doptype == 'p':
            Nidop = Ion(temp=T, ni_author=ni_author).update_dopant_ionisation(
                N_dop=Ndop, nxc=0, impurity='boron', author=ionauthor)
            n0, p0 = CF.get_carriers(Nidop, 0, 0, temp=T)
        for j in range(0, EtRange.shape[0]):
            n1 = NI().update(temp=T, author=ni_author) * \
                np.exp(EtRange[j] / kb / T)
            p1 = NI().update(temp=T, author=ni_author) * \
                np.exp(-EtRange[j] / kb / T)
            C = np.divide(m, (m + b))
            if doptype == 'n':
                Q = (1 - p1 / n0 - C) / (C + n1 / n0)
            elif doptype == 'p':
                Q = (C + p1 / p0) / (1 - n1 / p0 - C)
            if Q < 0:
                Q = np.nan
            k[j] = Q * vth_h / vth_e
            if doptype == 'n':
                TauMinor[j] = b / (1 + n1 / n0 + p1 / n0 /
                                   Q) * vth_h / vth_h300
            elif doptype == 'p':
                TauMinor[j] = b / (1 + Q * n1 / p0 + p1 /
                                   p0) * vth_e / vth_e300
        return EtRange, k, TauMinor

    def NT(self):
        for action in self.ionmodel.actions():
            if action.isChecked():
                ionauthor = action.text()
        for action in self.nimodel.actions():
            if action.isChecked():
                ni_author = action.text()
        for action in self.themodel.actions():
            if action.isChecked():
                theauthor = action.text()
        self.plotDPSS()
        mlist = []
        blist = []
        doplist = []
        Tlist = []
        doptypelist = []
        self.ntlist = []
        for item in self.listWidget_fitres.selectedItems():
            for fitres in self.fitreslist:
                if fitres.uid == item.data(32):
                    mlist.append(fitres.m)
                    blist.append(fitres.b)
                    doplist.append(fitres.Ndop)
                    Tlist.append(fitres.temp)
                    doptypelist.append(fitres.doptype)
        mb = nt.generatemb(m=mlist, b=blist)
        CaldtsList = nt.generteCaldts(
            T=Tlist, Ndop=doplist, doptypelist=doptypelist, ionauthor=ionauthor, vthauthor=theauthor, ni_author=ni_author)
        x0 = [0., 1e-16, 1]
        resupall = nt.NewtonMethodUP(x0=x0, mb=mb,
                                     CaldtsList=CaldtsList, MaxIntNum=2000)
        resdownall = nt.NewtonMethodDOWN(x0=x0, mb=mb,
                                         CaldtsList=CaldtsList, MaxIntNum=2000)
        if resupall is not np.nan:
            self.ntlist.append(resupall.reshape(3,))
            self.ax3.plot([resupall[0]], [resupall[1]], 'ro')
            self.ax4.plot([resupall[0]], [resupall[2]], 'ro')
        if resdownall is not np.nan:
            self.ntlist.append(resdownall.reshape(3,))
            self.ax3.plot([resdownall[0]], [resdownall[1]], 'ro')
            self.ax4.plot([resdownall[0]], [resdownall[2]], 'ro')
        if len(self.listWidget_fitres.selectedItems()) > 2:
            self.ntlist.append(np.asarray([0, 0, 0]))
            for i in range(len(self.listWidget_fitres.selectedItems()) - 1):
                for j in range(i + 1, len(self.listWidget_fitres.selectedItems())):
                    m2list = [mlist[i], mlist[j]]
                    b2list = [blist[i], blist[j]]
                    dop2list = [doplist[i], doplist[j]]
                    T2list = [Tlist[i], Tlist[j]]
                    doptype2list = [doptypelist[i], doptypelist[j]]
                    mb2 = nt.generatemb(m=m2list, b=b2list)
                    CaldtsList2 = nt.generteCaldts(
                        T=T2list, Ndop=dop2list, doptypelist=doptype2list)
                    x0 = [0., 1e-16, 1]
                    resup2 = nt.NewtonMethodUP(x0=x0, mb=mb2,
                                               CaldtsList=CaldtsList2, MaxIntNum=2000)
                    resdown2 = nt.NewtonMethodDOWN(x0=x0, mb=mb2,
                                                   CaldtsList=CaldtsList2, MaxIntNum=2000)
                    if resup2 is not np.nan:
                        self.ntlist.append(resup2.reshape(3,))
                        self.ax3.plot([resup2[0]], [resup2[1]], 'ko')
                        self.ax4.plot([resup2[0]], [resup2[2]], 'ko')
                    if resdown2 is not np.nan:
                        self.ntlist.append(resdown2.reshape(3,))
                        self.ax3.plot([resdown2[0]], [resdown2[1]], 'ko')
                        self.ax4.plot([resdown2[0]], [resdown2[2]], 'ko')
        self.canvas2.draw()

    def exportind(self):
        path = QFileDialog.getExistingDirectory(
            parent=self, caption='Select path for data export')
        fdpssname = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_res_DPSS.csv')
        fntname = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_res_NT.csv')
        alllist = [self.Etlist] + self.klist + self.tauminorlist + [self.stdk]
        alllist = np.asarray(alllist).T
        np.savetxt(fdpssname, alllist, delimiter=',',
                   header=self.dpsshead1 + self.dpsshead2, comments='')
        if hasattr(self, 'ntlist'):
            ntlist = np.asarray(self.ntlist)
            np.savetxt(fntname, ntlist, delimiter=',',
                       header='Et,TauMinor,k', comments='')

    def exportsimu(self):
        path = QFileDialog.getExistingDirectory(
            parent=self, caption='Select path for data export')
        f2dresi = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_res_residualmap.csv')
        f2dtauminor = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_res_tauminormap.csv')
        np.savetxt(f2dresi, self.residualmap, delimiter=',')
        if self.taummap[0, 0].shape == ():
            np.savetxt(f2dtauminor, self.taummap, delimiter=',')
        else:
            for i in range(self.taummap[0, 0].shape[-1]):
                f2dtauminor = os.path.join(
                    path, str(tm.mktime(tm.localtime())) + '_res_tauminormap_' + str(i) + '.csv')
                np.savetxt(f2dtauminor, self.taummap[:, :, i], delimiter=',')

    def expplot2(self):
        path = QFileDialog.getExistingDirectory(
            parent=self, caption='Select path for data export')
        fname = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_lifetimeandfitplot.csv')
        data = {}
        for line in self.ax1.get_lines():
            data[line.get_label() + '_' + self.ax1.get_xlabel()
                 ] = line.get_xdata()
            data[line.get_label() + '_' + self.ax1.get_ylabel()
                 ] = line.get_ydata()
        df = pd.DataFrame.from_dict(
            data, orient='index').transpose().fillna('')
        df.to_csv(fname, index=False)

    def advancedcheck(self):
        if self.radioButton_advfit.isChecked() == True:
            self.pushButton_expadvfit.setEnabled(True)
            self.pushButton_loadadvfit.setEnabled(True)
        else:
            self.pushButton_expadvfit.setEnabled(False)
            self.pushButton_loadadvfit.setEnabled(False)

        if self.radioButton_advfit.isChecked() == True and len(self.listWidget_fit.selectedItems()) != 0:
            if self.checkBox_intrinsic.isChecked() or self.checkBox_j0.isChecked() or self.checkBox_def1.isChecked() or self.checkBox_def2.isChecked() or self.checkBox_def3.isChecked():
                self.pushButton_plotadvfit.setEnabled(True)
                if self.checkBox_j0.isChecked():
                    for item in self.listWidget_fit.selectedItems():
                        for data in self.Rawdat:
                            if data.uid == item.data(32) and data.thickness is None:
                                self.pushButton_plotadvfit.setEnabled(False)
                    try:
                        if float(self.lineEdit_j0.text()) < 0:
                            self.pushButton_plotadvfit.setEnabled(False)
                    except ValueError:
                        self.pushButton_plotadvfit.setEnabled(False)
                if self.checkBox_def1.isChecked():
                    try:
                        if float(self.lineEdit_d1Et.text()) < -0.6 or float(self.lineEdit_d1Et.text()) > 0.6 or float(self.lineEdit_d1k.text()) < 0 or float(self.lineEdit_d1tau.text()) < 0:
                            self.pushButton_plotadvfit.setEnabled(False)
                    except ValueError:
                        self.pushButton_plotadvfit.setEnabled(False)
                if self.checkBox_def2.isChecked():
                    try:
                        if float(self.lineEdit_d2Et.text()) < -0.6 or float(self.lineEdit_d2Et.text()) > 0.6 or float(self.lineEdit_d2k.text()) < 0 or float(self.lineEdit_d2tau.text()) < 0:
                            self.pushButton_plotadvfit.setEnabled(False)
                    except ValueError:
                        self.pushButton_plotadvfit.setEnabled(False)
                if self.checkBox_def3.isChecked():
                    try:
                        if float(self.lineEdit_d3Et1.text()) < -0.6 or float(self.lineEdit_d3Et1.text()) > 0.6 or float(self.lineEdit_d3Et2.text()) < -0.6 or float(self.lineEdit_d3Et2.text()) > 0.6 or float(self.lineEdit_d3k1.text()) < 0 or float(self.lineEdit_d3k2.text()) < 0 or float(self.lineEdit_d3tau1.text()) < 0 or float(self.lineEdit_d3tau2.text()) < 0:
                            self.pushButton_plotadvfit.setEnabled(False)
                    except ValueError:
                        self.pushButton_plotadvfit.setEnabled(False)
            else:
                self.pushButton_plotadvfit.setEnabled(False)
        else:
            self.pushButton_plotadvfit.setEnabled(False)

        if self.radioButton_advfit.isChecked() == True and len(self.listWidget_fit.selectedItems()) != 0:
            if self.checkBox_j0.isChecked() or self.checkBox_def1.isChecked() or self.checkBox_def2.isChecked() or self.checkBox_def3.isChecked():
                self.pushButton_advfit.setEnabled(True)
                if self.checkBox_j0.isChecked():
                    for item in self.listWidget_fit.selectedItems():
                        for data in self.Rawdat:
                            if data.uid == item.data(32) and data.thickness is None:
                                self.pushButton_advfit.setEnabled(False)
            else:
                self.pushButton_advfit.setEnabled(False)
        else:
            self.pushButton_advfit.setEnabled(False)

        if self.checkBox_j0.isChecked():
            if self.lineEdit_j0.text() != "":
                try:
                    if float(self.lineEdit_j0.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixj0.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_j0l.text() != "" or self.lineEdit_j0u.text() != "":
                try:
                    if float(self.lineEdit_j0u.text()) < float(self.lineEdit_j0l.text()) or float(self.lineEdit_j0l.text()) < 0 or float(self.lineEdit_j0u.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)

        if self.checkBox_def1.isChecked():
            if self.lineEdit_d1Et.text() != "":
                try:
                    if float(self.lineEdit_d1Et.text()) < -0.6 or float(self.lineEdit_d1Et.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd1Et.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d1k.text() != "":
                try:
                    if float(self.lineEdit_d1k.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd1k.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d1tau.text() != "":
                try:
                    if float(self.lineEdit_d1tau.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d1Etl.text() != "" or self.lineEdit_d1Etu.text() != "":
                try:
                    if float(self.lineEdit_d1Etu.text()) < float(self.lineEdit_d1Etl.text()) or float(self.lineEdit_d1Etl.text()) < -0.6 or float(self.lineEdit_d1Etu.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d1kl.text() != "" or self.lineEdit_d1ku.text() != "":
                try:
                    if float(self.lineEdit_d1ku.text()) < float(self.lineEdit_d1kl.text()) or float(self.lineEdit_d1kl.text()) < 0 or float(self.lineEdit_d1ku.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d1taul.text() != "" or self.lineEdit_d1tauu.text() != "":
                try:
                    if float(self.lineEdit_d1tauu.text()) < float(self.lineEdit_d1taul.text()) or float(self.lineEdit_d1taul.text()) < 0 or float(self.lineEdit_d1tauu.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)

        if self.checkBox_def2.isChecked():
            if self.lineEdit_d2Et.text() != "":
                try:
                    if float(self.lineEdit_d2Et.text()) < -0.6 or float(self.lineEdit_d2Et.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd2Et.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d2k.text() != "":
                try:
                    if float(self.lineEdit_d2k.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd2k.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d2tau.text() != "":
                try:
                    if float(self.lineEdit_d2tau.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d2Etl.text() != "" or self.lineEdit_d2Etu.text() != "":
                try:
                    if float(self.lineEdit_d2Etu.text()) < float(self.lineEdit_d2Etl.text()) or float(self.lineEdit_d2Etl.text()) < -0.6 or float(self.lineEdit_d2Etu.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d2kl.text() != "" or self.lineEdit_d2ku.text() != "":
                try:
                    if float(self.lineEdit_d2ku.text()) < float(self.lineEdit_d2kl.text()) or float(self.lineEdit_d2kl.text()) < 0 or float(self.lineEdit_d2ku.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d2taul.text() != "" or self.lineEdit_d2tauu.text() != "":
                try:
                    if float(self.lineEdit_d2tauu.text()) < float(self.lineEdit_d2taul.text()) or float(self.lineEdit_d2taul.text()) < 0 or float(self.lineEdit_d2tauu.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)

        if self.checkBox_def3.isChecked():
            if self.lineEdit_d3Et1.text() != "":
                try:
                    if float(self.lineEdit_d3Et1.text()) < -0.6 or float(self.lineEdit_d3Et1.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd3Et1.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3Et2.text() != "":
                try:
                    if float(self.lineEdit_d3Et2.text()) < -0.6 or float(self.lineEdit_d3Et2.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd3Et2.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3k1.text() != "":
                try:
                    if float(self.lineEdit_d3k1.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd3k1.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3k2.text() != "":
                try:
                    if float(self.lineEdit_d3k2.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            elif self.checkBox_fixd3k2.isChecked():
                self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3tau1.text() != "":
                try:
                    if float(self.lineEdit_d3tau1.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3tau2.text() != "":
                try:
                    if float(self.lineEdit_d3tau2.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3Et1l.text() != "" or self.lineEdit_d3Et1u.text() != "":
                try:
                    if float(self.lineEdit_d3Et1u.text()) < float(self.lineEdit_d3Et1l.text()) or float(self.lineEdit_d3Et1l.text()) < -0.6 or float(self.lineEdit_d3Et1u.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3Et2l.text() != "" or self.lineEdit_d3Et2u.text() != "":
                try:
                    if float(self.lineEdit_d3Et2u.text()) < float(self.lineEdit_d3Et2l.text()) or float(self.lineEdit_d3Et2l.text()) < -0.6 or float(self.lineEdit_d3Et2u.text()) > 0.6:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3k1l.text() != "" or self.lineEdit_d3k1u.text() != "":
                try:
                    if float(self.lineEdit_d3k1u.text()) < float(self.lineEdit_d3k1l.text()) or float(self.lineEdit_d3k1l.text()) < 0 or float(self.lineEdit_d3k1u.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3k2l.text() != "" or self.lineEdit_d3k2u.text() != "":
                try:
                    if float(self.lineEdit_d3k2u.text()) < float(self.lineEdit_d3k2l.text()) or float(self.lineEdit_d3k2l.text()) < 0 or float(self.lineEdit_d3k2u.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3tau1l.text() != "" or self.lineEdit_d3tau1u.text() != "":
                try:
                    if float(self.lineEdit_d3tau1u.text()) < float(self.lineEdit_d3tau1l.text()) or float(self.lineEdit_d3tau1l.text()) < 0 or float(self.lineEdit_d3tau1u.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)
            if self.lineEdit_d3tau2l.text() != "" or self.lineEdit_d3tau2u.text() != "":
                try:
                    if float(self.lineEdit_d3tau2u.text()) < float(self.lineEdit_d3tau2l.text()) or float(self.lineEdit_d3tau2l.text()) < 0 or float(self.lineEdit_d3tau2u.text()) < 0:
                        self.pushButton_advfit.setEnabled(False)
                except ValueError:
                    self.pushButton_advfit.setEnabled(False)

    def expadvfit(self):
        path = QFileDialog.getExistingDirectory(
            parent=self, caption='Select path for data export')
        fadv = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_res.advfit.csv')
        header = ['Intrin_chk', 'j0_chk', 'j0', 'j0_fix', 'd1_chk', 'd1Et', 'd1Et_fix', 'd1k', 'd1k_fix', 'd1tau', 'd2_chk', 'd2Et', 'd2Et_fix', 'd2k', 'd2k_fix',
                  'd2tau', 'd3_chk', 'd3Et1', 'd3Et1_fix', 'd3Et2', 'd3Et2_fix', 'd3k1', 'd3k1_fix', 'd3k2', 'd3k2_fix', 'd3tau1', 'd3tau2']
        datlist = ['1' if self.checkBox_intrinsic.isChecked(
        ) else '0', '1' if self.checkBox_j0.isChecked() else '0', self.lineEdit_j0.text(), '1' if self.checkBox_fixj0.isChecked() else '0', '1' if self.checkBox_def1.isChecked() else '0', self.lineEdit_d1Et.text(), '1' if self.checkBox_fixd1Et.isChecked() else '0', self.lineEdit_d1k.text(), '1' if self.checkBox_fixd1k.isChecked() else '0', self.lineEdit_d1tau.text(), '1' if self.checkBox_def2.isChecked() else '0', self.lineEdit_d2Et.text(), '1' if self.checkBox_fixd2Et.isChecked() else '0', self.lineEdit_d2k.text(), '1' if self.checkBox_fixd2k.isChecked() else '0', self.lineEdit_d2tau.text(), '1' if self.checkBox_def3.isChecked() else '0', self.lineEdit_d3Et1.text(), '1' if self.checkBox_fixd3Et1.isChecked() else '0', self.lineEdit_d3Et2.text(), '1' if self.checkBox_fixd3Et2.isChecked() else '0', self.lineEdit_d3k1.text(), '1' if self.checkBox_fixd3k1.isChecked() else '0', self.lineEdit_d3k2.text(), '1' if self.checkBox_fixd3k2.isChecked() else '0', self.lineEdit_d3tau1.text(), self.lineEdit_d3tau2.text()]
        df = pd.DataFrame({'col1': header, 'col2': datlist})
        df.to_csv(fadv, header=None, index=None)

    def loadadvfit(self):
        fname = QFileDialog.getOpenFileName(
            self, caption='Choose advanced fit result file', filter='*.advfit.csv')
        if fname[0] != '':
            data = pd.read_csv(fname[0], names=['header', 'data'])
            data = data.set_index('header')
            try:
                if data.loc['Intrin_chk', 'data'] == 0:
                    self.checkBox_intrinsic.setChecked(False)
                elif data.loc['Intrin_chk', 'data'] == 1:
                    self.checkBox_intrinsic.setChecked(True)
                if data.loc['j0_chk', 'data'] == 0:
                    self.checkBox_j0.setChecked(False)
                elif data.loc['j0_chk', 'data'] == 1:
                    self.checkBox_j0.setChecked(True)
                if data.loc['d1_chk', 'data'] == 0:
                    self.checkBox_def1.setChecked(False)
                elif data.loc['d1_chk', 'data'] == 1:
                    self.checkBox_def1.setChecked(True)
                if data.loc['d2_chk', 'data'] == 0:
                    self.checkBox_def2.setChecked(False)
                elif data.loc['d2_chk', 'data'] == 1:
                    self.checkBox_def2.setChecked(True)
                if data.loc['d3_chk', 'data'] == 0:
                    self.checkBox_def3.setChecked(False)
                elif data.loc['d3_chk', 'data'] == 1:
                    self.checkBox_def3.setChecked(True)
                if data.loc['j0_fix', 'data'] == 0:
                    self.checkBox_fixj0.setChecked(False)
                elif data.loc['j0_fix', 'data'] == 1:
                    self.checkBox_fixj0.setChecked(True)
                if data.loc['d1Et_fix', 'data'] == 0:
                    self.checkBox_fixd1Et.setChecked(False)
                elif data.loc['d1Et_fix', 'data'] == 1:
                    self.checkBox_fixd1Et.setChecked(True)
                if data.loc['d1k_fix', 'data'] == 0:
                    self.checkBox_fixd1k.setChecked(False)
                elif data.loc['d1k_fix', 'data'] == 1:
                    self.checkBox_fixd1k.setChecked(True)
                if data.loc['d2Et_fix', 'data'] == 0:
                    self.checkBox_fixd2Et.setChecked(False)
                elif data.loc['d2Et_fix', 'data'] == 1:
                    self.checkBox_fixd2Et.setChecked(True)
                if data.loc['d2k_fix', 'data'] == 0:
                    self.checkBox_fixd2k.setChecked(False)
                elif data.loc['d2k_fix', 'data'] == 1:
                    self.checkBox_fixd2k.setChecked(True)
                if data.loc['d3Et1_fix', 'data'] == 0:
                    self.checkBox_fixd3Et1.setChecked(False)
                elif data.loc['d3Et1_fix', 'data'] == 1:
                    self.checkBox_fixd3Et1.setChecked(True)
                if data.loc['d3k1_fix', 'data'] == 0:
                    self.checkBox_fixd3k1.setChecked(False)
                elif data.loc['d3k1_fix', 'data'] == 1:
                    self.checkBox_fixd3k1.setChecked(True)
                if data.loc['d3Et2_fix', 'data'] == 0:
                    self.checkBox_fixd3Et2.setChecked(False)
                elif data.loc['d3Et2_fix', 'data'] == 1:
                    self.checkBox_fixd3Et2.setChecked(True)
                if data.loc['d3k2_fix', 'data'] == 0:
                    self.checkBox_fixd3k2.setChecked(False)
                elif data.loc['d3k2_fix', 'data'] == 1:
                    self.checkBox_fixd3k2.setChecked(True)
                if np.isnan(data.loc['j0', 'data']):
                    self.lineEdit_j0.clear()
                else:
                    self.lineEdit_j0.setText(
                        '{:.4e}'.format(data.loc['j0', 'data']))
                if np.isnan(data.loc['d1Et', 'data']):
                    self.lineEdit_d1Et.clear()
                else:
                    self.lineEdit_d1Et.setText(
                        '{:.3f}'.format(data.loc['d1Et', 'data']))
                if np.isnan(data.loc['d1k', 'data']):
                    self.lineEdit_d1k.clear()
                else:
                    self.lineEdit_d1k.setText(
                        '{:.3f}'.format(data.loc['d1k', 'data']))
                if np.isnan(data.loc['d1tau', 'data']):
                    self.lineEdit_d1tau.clear()
                else:
                    self.lineEdit_d1tau.setText(
                        '{:.3e}'.format(data.loc['d1tau', 'data']))
                if np.isnan(data.loc['d2Et', 'data']):
                    self.lineEdit_d2Et.clear()
                else:
                    self.lineEdit_d2Et.setText(
                        '{:.3f}'.format(data.loc['d2Et', 'data']))
                if np.isnan(data.loc['d2k', 'data']):
                    self.lineEdit_d2k.clear()
                else:
                    self.lineEdit_d2k.setText(
                        '{:.3f}'.format(data.loc['d2k', 'data']))
                if np.isnan(data.loc['d2tau', 'data']):
                    self.lineEdit_d2tau.clear()
                else:
                    self.lineEdit_d2tau.setText(
                        '{:.3e}'.format(data.loc['d2tau', 'data']))
                if np.isnan(data.loc['d3Et1', 'data']):
                    self.lineEdit_d3Et1.clear()
                else:
                    self.lineEdit_d3Et1.setText(
                        '{:.3f}'.format(data.loc['d3Et1', 'data']))
                if np.isnan(data.loc['d3k1', 'data']):
                    self.lineEdit_d3k1.clear()
                else:
                    self.lineEdit_d3k1.setText(
                        '{:.3f}'.format(data.loc['d3k1', 'data']))
                if np.isnan(data.loc['d3tau1', 'data']):
                    self.lineEdit_d3tau1.clear()
                else:
                    self.lineEdit_d3tau1.setText(
                        '{:.3e}'.format(data.loc['d3tau1', 'data']))
                if np.isnan(data.loc['d3Et2', 'data']):
                    self.lineEdit_d3Et2.clear()
                else:
                    self.lineEdit_d3Et2.setText(
                        '{:.2e}'.format(data.loc['d3Et2', 'data']))
                if np.isnan(data.loc['d3k2', 'data']):
                    self.lineEdit_d3k2.clear()
                else:
                    self.lineEdit_d3k2.setText(
                        '{:.2e}'.format(data.loc['d3k2', 'data']))
                if np.isnan(data.loc['d3tau2', 'data']):
                    self.lineEdit_d3tau2.clear()
                else:
                    self.lineEdit_d3tau2.setText(
                        '{:.3e}'.format(data.loc['d3tau2', 'data']))
            except:
                pass

    def loaddata(self):
        try:
            filename = QFileDialog.getOpenFileNames(
                self, caption='Choose data file', filter="Exported fake Sinton file(*.xlsm)")
            if filename[0] != '':
                for fname in filename[0]:
                    nxc = []
                    tau = []
                    wb = xls.load_workbook(filename=fname, data_only=True)
                    ws = wb['RawData']
                    lenth = ws.max_row
                    for row in ws.iter_rows('E5:G' + str(lenth)):
                        if row[0].value is not None and row[-1].value is not None:
                            tau.append(row[0].value)
                            nxc.append(row[-1].value)
                    nxc = np.asarray(nxc)
                    tau = np.asarray(tau)
                    idx = np.argsort(nxc)
                    nxc = nxc[idx]
                    tau = tau[idx]
                    ws1 = wb['User']
                    Ndop = ws1['J9'].value
                    doptype = ws1['D6'].value[0]
                    thickness = ws1['B6'].value
                    j0 = ws1['D9'].value
                    if ws1['L8'].value[0] == 'T':
                        temp = ws1['L9'].value + 273.15
                    else:
                        temp = 303
                    Sinton = rawdata(
                        name='Sinton_' + (os.path.splitext(os.path.basename(fname))[0]), nxc=nxc, tau=tau, PCPL='PC', Ndop=Ndop, doptype=doptype, temp=temp, thickness=thickness, j0=j0, uid=1 + self.getmaxuid())
                    self.Rawdat.append(Sinton)
                    dataitem = QListWidgetItem(parent=self.listWidget_fit)
                    dataitem.setText(Sinton.name)
                    dataitem.setData(32, Sinton.uid)
                    self.listWidget_fit.addItem(dataitem)
        except Exception as e:
            print(e)
            print('You probably have modified the original file in a unaaceptable way')

    def getmaxuid(self):
        a = -1
        for i in range(self.listWidget_fit.count()):
            if self.listWidget_fit.item(i).data(32) > a:
                a = self.listWidget_fit.item(i).data(32)
        return a

    def generatefitopt(self):
        fitopt = []
        fitopt2 = []
        if self.checkBox_intrinsic.isChecked():
            fitopt2.append(1)
        else:
            fitopt2.append(0)
        if self.checkBox_j0.isChecked():
            fitopt2.append(1)
            if not self.checkBox_fixj0.isChecked():
                fitopt.append(1)
            else:
                fitopt.append(0)
        else:
            fitopt.append(0)
            fitopt2.append(0)
        if self.checkBox_def1.isChecked():
            fitopt2.append(1)
            if self.checkBox_fixd1Et.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            if self.checkBox_fixd1k.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            fitopt.append(1)
        else:
            fitopt.append(0)
            fitopt.append(0)
            fitopt.append(0)
            fitopt2.append(0)
        if self.checkBox_def2.isChecked():
            fitopt2.append(1)
            if self.checkBox_fixd2Et.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            if self.checkBox_fixd2k.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            fitopt.append(1)
        else:
            fitopt.append(0)
            fitopt.append(0)
            fitopt.append(0)
            fitopt2.append(0)
        if self.checkBox_def3.isChecked():
            fitopt2.append(1)
            if self.checkBox_fixd3Et1.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            if self.checkBox_fixd3Et2.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            if self.checkBox_fixd3k1.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            if self.checkBox_fixd3k2.isChecked():
                fitopt.append(0)
            else:
                fitopt.append(1)
            fitopt.append(1)
            fitopt.append(1)
        else:
            fitopt.append(0)
            fitopt.append(0)
            fitopt.append(0)
            fitopt.append(0)
            fitopt.append(0)
            fitopt.append(0)
            fitopt2.append(0)
        return fitopt, fitopt2

    def generatex0list(self):
        x0list = []
        if self.lineEdit_j0.text() != "":
            try:
                x0list.append(float(self.lineEdit_j0.text()))
            except:
                x0list.append(0)
                self.lineEdit_j0.clear()
        else:
            x0list.append(0)
        if self.lineEdit_d1Et.text() != "":
            try:
                x0list.append(float(self.lineEdit_d1Et.text()))
            except:
                x0list.append(0)
                self.lineEdit_d1Et.clear()
        else:
            x0list.append(0)
        if self.lineEdit_d1k.text() != "":
            try:
                x0list.append(float(self.lineEdit_d1k.text()))
            except:
                x0list.append(1)
                self.lineEdit_d1k.clear()
        else:
            x0list.append(1)
        if self.lineEdit_d1tau.text() != "":
            try:
                x0list.append(float(self.lineEdit_d1tau.text()))
            except:
                x0list.append(1e-6)
                self.lineEdit_d1tau.clear()
        else:
            x0list.append(1e-6)
        if self.lineEdit_d2Et.text() != "":
            try:
                x0list.append(float(self.lineEdit_d2Et.text()))
            except:
                x0list.append(0)
                self.lineEdit_d2Et.clear()
        else:
            x0list.append(0)
        if self.lineEdit_d2k.text() != "":
            try:
                x0list.append(float(self.lineEdit_d2k.text()))
            except:
                x0list.append(1)
                self.lineEdit_d2k.clear()
        else:
            x0list.append(1)
        if self.lineEdit_d2tau.text() != "":
            try:
                x0list.append(float(self.lineEdit_d2tau.text()))
            except:
                x0list.append(1e-6)
                self.lineEdit_d2tau.clear()
        else:
            x0list.append(1e-6)
        if self.lineEdit_d3Et1.text() != "":
            try:
                x0list.append(float(self.lineEdit_d3Et1.text()))
            except:
                x0list.append(0)
                self.lineEdit_d3Et1.clear()
        else:
            x0list.append(0)
        if self.lineEdit_d3Et2.text() != "":
            try:
                x0list.append(float(self.lineEdit_d3Et2.text()))
            except:
                x0list.append(0)
                self.lineEdit_d3Et2.clear()
        else:
            x0list.append(0)
        if self.lineEdit_d3k1.text() != "":
            try:
                x0list.append(float(self.lineEdit_d3k1.text()))
            except:
                x0list.append(1)
                self.lineEdit_d3k1.clear()
        else:
            x0list.append(1)
        if self.lineEdit_d3k2.text() != "":
            try:
                x0list.append(float(self.lineEdit_d3k2.text()))
            except:
                x0list.append(1)
                self.lineEdit_d3k2.clear()
        else:
            x0list.append(1)
        if self.lineEdit_d3tau1.text() != "":
            try:
                x0list.append(float(self.lineEdit_d3tau1.text()))
            except:
                x0list.append(1e-6)
                self.lineEdit_d3tau1.clear()
        else:
            x0list.append(1e-6)
        if self.lineEdit_d3tau2.text() != "":
            try:
                x0list.append(float(self.lineEdit_d3tau2.text()))
            except:
                x0list.append(1e-6)
                self.lineEdit_d3tau2.clear()
        else:
            x0list.append(1e-6)
        fitoptlist, fitoptlist2 = self.generatefitopt()
        x0nlist = []
        for opt, x0 in zip(fitoptlist, x0list):
            if opt == 1:
                x0nlist.append(x0)
        return x0nlist, x0list

    def generatebounds(self):
        bounds = []
        j0default = (1e-24, None)
        Etdefault = (-0.6, 0.6)
        kdefault = (1e-4, 10000)
        taudefault = (0, None)
        if self.lineEdit_j0l.text() != "":
            try:
                j0l = float(self.lineEdit_j0l.text())
            except:
                j0l = j0default[0]
                self.lineEdit_j0l.clear()
        if self.lineEdit_j0u.text() != "":
            try:
                j0u = float(self.lineEdit_j0u.text())
            except:
                j0u = j0default[1]
                self.lineEdit_j0u.clear()
        bounds.append((j0l, j0u))
        if self.lineEdit_d1Etl.text() != "":
            try:
                Etl = (float(self.lineEdit_d1Etl.text()))
            except:
                Etl = Etdefault[0]
                self.lineEdit_d1Etl.clear()
        if self.lineEdit_d1Etu.text() != "":
            try:
                Etu = (float(self.lineEdit_d1Etu.text()))
            except:
                Etu = Etdefault[1]
                self.lineEdit_d1Etu.clear()
        bounds.append((Etl, Etu))
        if self.lineEdit_d1kl.text() != "":
            try:
                kl = (float(self.lineEdit_d1kl.text()))
            except:
                kl = kdefault[0]
                self.lineEdit_d1kl.clear()
        if self.lineEdit_d1ku.text() != "":
            try:
                ku = (float(self.lineEdit_d1ku.text()))
            except:
                ku = kdefault[1]
                self.lineEdit_d1ku.clear()
        bounds.append((kl, ku))
        if self.lineEdit_d1taul.text() != "":
            try:
                taubl = (float(self.lineEdit_d1taul.text()))
            except:
                taubl = taudefault[0]
                self.lineEdit_d1taul.clear()
        if self.lineEdit_d1tauu.text() != "":
            try:
                taubu = (float(self.lineEdit_d1tauu.text()))
            except:
                taubu = taudefault[1]
                self.lineEdit_d1tauu.clear()
        bounds.append((taubl, taubu))
        if self.lineEdit_d2Etl.text() != "":
            try:
                Etl = (float(self.lineEdit_d2Etl.text()))
            except:
                Etl = Etdefault[0]
                self.lineEdit_d2Etl.clear()
        if self.lineEdit_d2Etu.text() != "":
            try:
                Etu = (float(self.lineEdit_d2Etu.text()))
            except:
                Etu = Etdefault[1]
                self.lineEdit_d2Etu.clear()
        bounds.append((Etl, Etu))
        if self.lineEdit_d2kl.text() != "":
            try:
                kl = (float(self.lineEdit_d2kl.text()))
            except:
                kl = kdefault[0]
                self.lineEdit_d2kl.clear()
        if self.lineEdit_d2ku.text() != "":
            try:
                ku = (float(self.lineEdit_d2ku.text()))
            except:
                ku = kdefault[1]
                self.lineEdit_d2ku.clear()
        bounds.append((kl, ku))
        if self.lineEdit_d2taul.text() != "":
            try:
                taubl = (float(self.lineEdit_d2taul.text()))
            except:
                taubl = taudefault[0]
                self.lineEdit_d2taul.clear()
        if self.lineEdit_d2tauu.text() != "":
            try:
                taubu = (float(self.lineEdit_d2tauu.text()))
            except:
                taubu = taudefault[1]
                self.lineEdit_d2tauu.clear()
        bounds.append((taubl, taubu))
        if self.lineEdit_d3Et1l.text() != "":
            try:
                Etl = (float(self.lineEdit_d3Et1l.text()))
            except:
                Etl = Etdefault[0]
                self.lineEdit_d3Et1l.clear()
        if self.lineEdit_d3Et1u.text() != "":
            try:
                Etu = (float(self.lineEdit_d3Et1u.text()))
            except:
                Etu = Etdefault[1]
                self.lineEdit_d3Et1u.clear()
        bounds.append((Etl, Etu))
        if self.lineEdit_d3Et2l.text() != "":
            try:
                Etl = (float(self.lineEdit_d3Et2l.text()))
            except:
                Etl = Etdefault[0]
                self.lineEdit_d3Et2l.clear()
        if self.lineEdit_d3Et2u.text() != "":
            try:
                Etu = (float(self.lineEdit_d3Et2u.text()))
            except:
                Etu = Etdefault[1]
                self.lineEdit_d3Et2u.clear()
        bounds.append((Etl, Etu))
        if self.lineEdit_d3k1l.text() != "":
            try:
                kl = (float(self.lineEdit_d3k1l.text()))
            except:
                kl = kdefault[0]
                self.lineEdit_d3k1l.clear()
        if self.lineEdit_d3k1u.text() != "":
            try:
                ku = (float(self.lineEdit_d3k1u.text()))
            except:
                ku = kdefault[1]
                self.lineEdit_d3k1u.clear()
        bounds.append((kl, ku))
        if self.lineEdit_d3k2l.text() != "":
            try:
                kl = (float(self.lineEdit_d3k2l.text()))
            except:
                kl = kdefault[0]
                self.lineEdit_d3k2l.clear()
        if self.lineEdit_d3k2u.text() != "":
            try:
                ku = (float(self.lineEdit_d3k2u.text()))
            except:
                ku = kdefault[1]
                self.lineEdit_d3k2u.clear()
        bounds.append((kl, ku))
        if self.lineEdit_d3tau1l.text() != "":
            try:
                taubl = (float(self.lineEdit_d3tau1l.text()))
            except:
                taubl = taudefault[0]
                self.lineEdit_d3tau1l.clear()
        if self.lineEdit_d3tau1u.text() != "":
            try:
                taubu = (float(self.lineEdit_d3tau1u.text()))
            except:
                taubu = taudefault[1]
                self.lineEdit_d3tau1u.clear()
        bounds.append((taubl, taubu))
        if self.lineEdit_d3tau2l.text() != "":
            try:
                taubl = (float(self.lineEdit_d3tau2l.text()))
            except:
                taubl = taudefault[0]
                self.lineEdit_d3tau2l.clear()
        if self.lineEdit_d3tau2u.text() != "":
            try:
                taubu = (float(self.lineEdit_d3tau2u.text()))
            except:
                taubu = taudefault[1]
                self.lineEdit_d3tau2u.clear()
        bounds.append((taubl, taubu))
        if self.checkBox_j0.isChecked():
            if self.checkBox_fixj0.isChecked():
                bounds[0] = (float(self.lineEdit_j0.text()),
                             float(self.lineEdit_j0.text()))
        else:
            bounds[0] = (0.1, 0.1)
        if self.checkBox_def1.isChecked():
            if self.checkBox_fixd1Et.isChecked():
                bounds[1] = (float(self.lineEdit_d1Et.text()),
                             float(self.lineEdit_d1Et.text()))
            if self.checkBox_fixd1k.isChecked():
                bounds[2] = (float(self.lineEdit_d1k.text()),
                             float(self.lineEdit_d1k.text()))
        else:
            bounds[1] = (0, 0)
            bounds[2] = (1, 1)
            bounds[3] = (1e-6, 1e-6)
        if self.checkBox_def2.isChecked():
            if self.checkBox_fixd2Et.isChecked():
                bounds[4] = (float(self.lineEdit_d2Et.text()),
                             float(self.lineEdit_d2Et.text()))
            if self.checkBox_fixd1k.isChecked():
                bounds[5] = (float(self.lineEdit_d2k.text()),
                             float(self.lineEdit_d2k.text()))
        else:
            bounds[4] = (0, 0)
            bounds[5] = (1, 1)
            bounds[6] = (1e-6, 1e-6)
        if self.checkBox_def3.isChecked():
            if self.checkBox_fixd3Et1.isChecked():
                bounds[7] = (float(self.lineEdit_d3Et1.text()),
                             float(self.lineEdit_d3Et1.text()))
            if self.checkBox_fixd3Et2.isChecked():
                bounds[8] = (float(self.lineEdit_d3Et2.text()),
                             float(self.lineEdit_d3Et2.text()))
            if self.checkBox_fixd3k1.isChecked():
                bounds[9] = (float(self.lineEdit_d3k1.text()),
                             float(self.lineEdit_d3k1.text()))
            if self.checkBox_fixd3k2.isChecked():
                bounds[10] = (float(self.lineEdit_d3k2.text()),
                              float(self.lineEdit_d3k2.text()))
        else:
            bounds[7] = (0, 0)
            bounds[8] = (0, 0)
            bounds[9] = (1, 1)
            bounds[10] = (1, 1)
            bounds[11] = (1e-6, 1e-6)
            bounds[12] = (1e-6, 1e-6)
        boundlist = []
        fitoptlist, fitoptlist2 = self.generatefitopt()
        for opt, bd in zip(fitoptlist, bounds):
            if opt == 1:
                boundlist.append(bd)
        return boundlist

    def generatedatalist(self):
        nxclist = []
        taulist = []
        Ndoplist = []
        doptypelist = []
        Tlist = []
        thicklist = []
        for item in self.listWidget_fit.selectedItems():
            for data in self.Rawdat:
                if data.uid == item.data(32):
                    nxclist.append(data.nxc)
                    taulist.append(data.tau)
                    Ndoplist.append(data.Ndop)
                    doptypelist.append(data.doptype)
                    Tlist.append(data.temp)
                    thicklist.append(data.thickness)
        for action in self.ionmodel.actions():
            if action.isChecked():
                ionauthor = action.text()
        for action in self.nimodel.actions():
            if action.isChecked():
                ni_author = action.text()
        for action in self.themodel.actions():
            if action.isChecked():
                theauthor = action.text()
        for action in self.Radmodel.actions():
            if action.isChecked():
                rad_author = action.text()
        for action in self.Augmodel.actions():
            if action.isChecked():
                aug_author = action.text()
        authorlist = [ionauthor, ni_author, theauthor, rad_author, aug_author]
        return nxclist, taulist, Ndoplist, doptypelist, Tlist, thicklist, authorlist

    def advsametaum(self, fitobj, nxclist, taulist, Ndoplist, doptypelist, Tlist, thicklist, authorlist, fitopt, fitopt2, totx0list, **kwarg):
        nxclist = np.asarray(nxclist)
        taulist = np.asarray(taulist) * 1e6
        inv_int = taulist * 0
        inv_j0 = taulist * 0
        inv_def1 = taulist * 0
        inv_def2 = taulist * 0
        inv_def3 = taulist * 0
        Nalist = []
        Ndlist = []
        for nxc, Ndop, doptype, T in zip(nxclist, Ndoplist, doptypelist, Tlist):
            if doptype == 'n':
                Ndlist.append(Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                    author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='phosphorous'))
                Nalist.append(1)
            if doptype == 'p':
                Nalist.append(Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                    author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='boron'))
                Ndlist.append(1)
        if fitopt2[0] == 1:
            inv_int = []
            for nxc, Na, Nd, T in zip(nxclist, Nalist, Ndlist, Tlist):
                inv_int.append(rad().itau(ni_author=authorlist[1], author=authorlist[3], temp=T, Na=Na, Nd=Nd, nxc=nxc) +
                               aug().itau(ni_author=authorlist[1], author=authorlist[4],
                                          temp=T, Na=Na, Nd=Nd, nxc=nxc))
            inv_int = np.asarray(inv_int)
        if fitopt2[1] == 1:
            inv_j0 = []
            if fitopt[0] == 1:
                j0 = fitobj[0]
            else:
                j0 = totx0list[0]
            for nxc, Ndop, doptypelist, T, thick in zip(nxclist, Ndoplist, doptypelist, Tlist, thicklist):
                if thick == None:
                    thick = 100
                ni = NI(temp=T).update(author=authorlist[1])
                if doptype == 'n':
                    iNdop = Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                        author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='phosphorous')
                if doptype == 'p':
                    iNdop = Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                        author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='boron')
                inv_j0.append(2 * j0 * (iNdop + nxc) / const.e / thick / ni**2)
            inv_j0 = np.asarray(inv_j0)
        if fitopt2[2] == 1:
            if fitopt[1] == 1:
                d1Et = fitobj[sum(fitopt[:1])]
            else:
                d1Et = totx0list[1]
            if fitopt[2] == 1:
                d1k = fitobj[sum(fitopt[:2])]
            else:
                d1k = totx0list[2]
            if fitopt[3] == 1:
                d1tau = fitobj[sum(fitopt[:3])]
            else:
                d1tau = totx0list[3]
            inv_def1 = []
            for nxc, Ndop, doptype, T in zip(nxclist, Ndoplist, doptypelist, Tlist):
                inv_def1.append(1 / self.SRHtau2(nxc, d1Et, d1tau, d1k, T, Ndop,
                                                 doptype, authorlist[1], authorlist[2], authorlist[0]))
            inv_def1 = np.asarray(inv_def1)
        if fitopt2[3] == 1:
            if fitopt[4] == 1:
                d2Et = fitobj[sum(fitopt[:4])]
            else:
                d2Et = totx0list[4]
            if fitopt[5] == 1:
                d2k = fitobj[sum(fitopt[:5])]
            else:
                d2k = totx0list[5]
            if fitopt[6] == 1:
                d2tau = fitobj[sum(fitopt[:6])]
            else:
                d2tau = totx0list[6]
            inv_def2 = []
            for nxc, Ndop, doptype, T in zip(nxclist, Ndoplist, doptypelist, Tlist):
                inv_def2.append(1 / self.SRHtau2(nxc, d2Et, d2tau, d2k, T, Ndop,
                                                 doptype, authorlist[1], authorlist[2], authorlist[0]))
            inv_def2 = np.asarray(inv_def2)
        if fitopt2[4] == 1:
            if fitopt[7] == 1:
                d3Et1 = fitobj[sum(fitopt[:7])]
            else:
                d3Et1 = totx0list[7]
            if fitopt[8] == 1:
                d3Et2 = fitobj[sum(fitopt[:8])]
            else:
                d3Et2 = totx0list[8]
            if fitopt[9] == 1:
                d3k1 = fitobj[sum(fitopt[:9])]
            else:
                d3k1 = totx0list[9]
            if fitopt[10] == 1:
                d3k2 = fitobj[sum(fitopt[:10])]
            else:
                d3k2 = totx0list[10]
            if fitopt[11] == 1:
                d3tau1 = fitobj[sum(fitopt[:11])]
            else:
                d3tau1 = totx0list[11]
            if fitopt[12] == 1:
                d3tau2 = fitobj[sum(fitopt[:12])]
            else:
                d3tau2 = totx0list[12]
            inv_def3 = []
            for nxc, Ndop, doptype, T in zip(nxclist, Ndoplist, doptypelist, Tlist):
                inv_def3.append(1 / self.SRHtau3(nxc, d3Et1, d3Et2, d3tau1, d3tau2, d3k1,
                                                 d3k2, T, Ndop, doptype, authorlist[1], authorlist[2], authorlist[0]))
            inv_def3 = np.asarray(inv_def3)
        tottau = 1e6 / (inv_int + inv_j0 + inv_def1 + inv_def2 + inv_def3)
        squarederror = np.sum((tottau - np.asarray(taulist))
                              * (tottau - np.asarray(taulist)))
        return squarederror

    def advdifftaum(self, fitobj, nxclist, taulist, Ndoplist, doptypelist, Tlist, thicklist, authorlist, **kwarg):
        pass

    def advfit(self):
        fitopt, fitopt2 = self.generatefitopt()
        x0list, x0totlist = self.generatex0list()
        bounds = self.generatebounds()
        nxclist, taulist, Ndoplist, doptypelist, Tlist, thicklist, authorlist = self.generatedatalist()
        # print(fitopt)
        # print(fitopt2)
        # print(x0list)
        # print(bounds)
        # print(self.advsametaum(x0list, nxclist, taulist, Ndoplist,
        # doptypelist, Tlist, thicklist, authorlist, fitopt, fitopt2,
        # x0totlist))
        if self.checkBox_sametaum_adv.isChecked():
            fitres = minimize(self.advsametaum, x0=x0list, args=(
                nxclist, taulist, Ndoplist, doptypelist, Tlist, thicklist, authorlist, fitopt, fitopt2, x0totlist), bounds=bounds)
        else:
            fitres = minimize(self.advdifftaum, x0=x0list, args=(
                nxclist, taulist, Ndoplist, doptypelist, Tlist, thicklist, authorlist, fitopt, fitopt2, x0totlist), bounds=bounds)
        # print(fitres)
        if fitopt[0] == 1:
            self.lineEdit_j0.setText('{:.4e}'.format(fitres.x[0]))
        if fitopt[1] == 1:
            self.lineEdit_d1Et.setText(
                '{:.3f}'.format(fitres[sum(fitopt[:1])]))
        if fitopt[2] == 1:
            self.lineEdit_d1k.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:2])]))
        if fitopt[3] == 1:
            self.lineEdit_d1tau.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:3])]))
        if fitopt[4] == 1:
            self.lineEdit_d2Et.setText(
                '{:.3f}'.format(fitres.x[sum(fitopt[:4])]))
        if fitopt[5] == 1:
            self.lineEdit_d2k.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:5])]))
        if fitopt[6] == 1:
            self.lineEdit_d2tau.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:6])]))
        if fitopt[7] == 1:
            self.lineEdit_d3Et1.setText(
                '{:.3f}'.format(fitres.x[sum(fitopt[:7])]))
        if fitopt[8] == 1:
            self.lineEdit_d3Et2.setText(
                '{:.3f}'.format(fitres.x[sum(fitopt[:8])]))
        if fitopt[9] == 1:
            self.lineEdit_d3k1.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:9])]))
        if fitopt[10] == 1:
            self.lineEdit_d3k2.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:10])]))
        if fitopt[11] == 1:
            self.lineEdit_d3tau1.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:11])]))
        if fitopt[12] == 1:
            self.lineEdit_d3tau2.setText(
                '{:.3e}'.format(fitres.x[sum(fitopt[:12])]))

    def calculatelifetimes(self, x0list, nxc, tau, Ndop, doptype, T, thick, authorlist, fitopt, fitopt2, totx0list):
        nxc = np.asarray(nxc)
        tau = np.asarray(tau) * 1e6
        inv_int = tau * 0
        inv_j0 = tau * 0
        inv_def1 = tau * 0
        inv_def2 = tau * 0
        inv_def3 = tau * 0
        if doptype == 'n':
            Nd = Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='phosphorous')
            Na = 1
        if doptype == 'p':
            Na = Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='boron')
            Nd = 1
        if fitopt2[0] == 1:
            inv_int = np.asarray(rad().itau(ni_author=authorlist[1], author=authorlist[3], temp=T, Na=Na, Nd=Nd, nxc=nxc) +
                                 aug().itau(ni_author=authorlist[1], author=authorlist[4],
                                            temp=T, Na=Na, Nd=Nd, nxc=nxc))
        if fitopt2[1] == 1:
            if fitopt[0] == 1:
                j0 = x0list[0]
            else:
                j0 = totx0list[0]
            if thick == None:
                thick = 100
            ni = NI(temp=T).update(author=authorlist[1])
            if doptype == 'n':
                iNdop = Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                    author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='phosphorous')
            if doptype == 'p':
                iNdop = Ion(temp=T, ni_author=authorlist[1]).update_dopant_ionisation(
                    author=authorlist[0], N_dop=Ndop, nxc=nxc, impurity='boron')
            inv_j0 = np.asarray(2 * j0 * (iNdop + nxc) /
                                const.e / thick / ni**2)
        if fitopt2[2] == 1:
            if fitopt[1] == 1:
                d1Et = x0list[sum(fitopt[:1])]
            else:
                d1Et = totx0list[1]
            if fitopt[2] == 1:
                d1k = x0list[sum(fitopt[:2])]
            else:
                d1k = totx0list[2]
            if fitopt[3] == 1:
                d1tau = x0list[sum(fitopt[:3])]
            else:
                d1tau = totx0list[3]
            inv_def1 = np.asarray(1 / self.SRHtau2(nxc, d1Et, d1tau, d1k, T, Ndop,
                                                   doptype, authorlist[1], authorlist[2], authorlist[0]))
        if fitopt2[3] == 1:
            if fitopt[4] == 1:
                d2Et = x0list[sum(fitopt[:4])]
            else:
                d2Et = totx0list[4]
            if fitopt[5] == 1:
                d2k = x0list[sum(fitopt[:5])]
            else:
                d2k = totx0list[5]
            if fitopt[6] == 1:
                d2tau = x0list[sum(fitopt[:6])]
            else:
                d2tau = totx0list[6]
            inv_def2 = np.asarray(1 / self.SRHtau2(nxc, d2Et, d2tau, d2k, T, Ndop,
                                                   doptype, authorlist[1], authorlist[2], authorlist[0]))
        if fitopt2[4] == 1:
            if fitopt[7] == 1:
                d3Et1 = x0list[sum(fitopt[:7])]
            else:
                d3Et1 = totx0list[7]
            if fitopt[8] == 1:
                d3Et2 = x0list[sum(fitopt[:8])]
            else:
                d3Et2 = totx0list[8]
            if fitopt[9] == 1:
                d3k1 = x0list[sum(fitopt[:9])]
            else:
                d3k1 = totx0list[9]
            if fitopt[10] == 1:
                d3k2 = x0list[sum(fitopt[:10])]
            else:
                d3k2 = totx0list[10]
            if fitopt[11] == 1:
                d3tau1 = x0list[sum(fitopt[:11])]
            else:
                d3tau1 = totx0list[11]
            if fitopt[12] == 1:
                d3tau2 = x0list[sum(fitopt[:12])]
            else:
                d3tau2 = totx0list[12]
            inv_def3 = np.asarray(1 / self.SRHtau3(nxc, d3Et1, d3Et2, d3tau1, d3tau2, d3k1,
                                                   d3k2, T, Ndop, doptype, authorlist[1], authorlist[2], authorlist[0]))
        tottau = 1 / (inv_int + inv_j0 + inv_def1 + inv_def2 + inv_def3)
        return tottau, inv_int, inv_j0, inv_def1, inv_def2, inv_def3

    def plotadvfit(self):
        fitopt, fitopt2 = self.generatefitopt()
        x0list, x0totlist = self.generatex0list()
        dummy, dummy, dummy, dummy, dummy, dummy, authorlist = self.generatedatalist()
        self.ax1.clear()
        self.ax1.grid()
        for item in self.listWidget_fit.selectedItems():
            for data in self.Rawdat:
                if data.uid == item.data(32):
                    tottau, inv_int, inv_j0, inv_def1, inv_def2, inv_def3 = self.calculatelifetimes(
                        x0list, data.nxc, data.tau, data.Ndop, data.doptype, data.temp, data.thickness, authorlist, fitopt, fitopt2, x0totlist)
                    if self.comboBox_plotopt2.currentIndex() == 0:
                        X = self.getXY(nxc=data.nxc, T=data.temp,
                                       Ndop=data.Ndop, doptype=data.doptype)
                        self.ax1.set_xlabel(r'X/Y')
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.plot(X, data.tau, '.', label=data.name)
                        self.ax1.plot(X, tottau, label=data.name + 'fit')
                        if self.checkBox_intrinsic.isChecked():
                            self.ax1.plot(X, 1 / inv_int,
                                          label=data.name + 'intinsic')
                        if self.checkBox_j0.isChecked():
                            self.ax1.plot(
                                X, 1 / inv_j0, label=data.name + 'j0')
                        if self.checkBox_def1.isChecked():
                            self.ax1.plot(
                                X, 1 / inv_def1, label=data.name + 'def1')
                        if self.checkBox_def2.isChecked():
                            self.ax1.plot(
                                X, 1 / inv_def2, label=data.name + 'def2')
                        if self.checkBox_def3.isChecked():
                            self.ax1.plot(
                                X, 1 / inv_def3, label=data.name + 'def3')
                    elif self.comboBox_plotopt2.currentIndex() == 1:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, data.tau, '.', label=data.name)
                        self.ax1.plot(data.nxc, tottau,
                                      label=data.name + 'fit')
                        if self.checkBox_intrinsic.isChecked():
                            self.ax1.plot(data.nxc, 1 / inv_int,
                                          label=data.name + 'intinsic')
                        if self.checkBox_j0.isChecked():
                            self.ax1.plot(
                                data.nxc, 1 / inv_j0, label=data.name + 'j0')
                        if self.checkBox_def1.isChecked():
                            self.ax1.plot(
                                data.nxc, 1 / inv_def1, label=data.name + 'def1')
                        if self.checkBox_def2.isChecked():
                            self.ax1.plot(
                                data.nxc, 1 / inv_def2, label=data.name + 'def2')
                        if self.checkBox_def3.isChecked():
                            self.ax1.plot(
                                data.nxc, 1 / inv_def3, label=data.name + 'def3')
                    elif self.comboBox_plotopt2.currentIndex() == 2:
                        self.ax1.semilogx()
                        self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                        self.ax1.set_xlabel(
                            r'Excess carrier density $[cm^{-3}]$')
                        self.ax1.plot(data.nxc, 1. / data.tau,
                                      '.', label=data.name)
                        self.ax1.plot(data.nxc, 1 / tottau,
                                      label=data.name + 'fit')
                        if self.checkBox_intrinsic.isChecked():
                            self.ax1.plot(data.nxc, inv_int,
                                          label=data.name + 'intinsic')
                        if self.checkBox_j0.isChecked():
                            self.ax1.plot(
                                data.nxc, inv_j0, label=data.name + 'j0')
                        if self.checkBox_def1.isChecked():
                            self.ax1.plot(
                                data.nxc, inv_def1, label=data.name + 'def1')
                        if self.checkBox_def2.isChecked():
                            self.ax1.plot(
                                data.nxc, inv_def2, label=data.name + 'def2')
                        if self.checkBox_def3.isChecked():
                            self.ax1.plot(
                                data.nxc, inv_def3, label=data.name + 'def3')
        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    LSana_dpss = LSana_dpss(rawdata=None, uidlist=None,
                            mainwindow=None)
    LSana_dpss.show()
    sys.exit(app.exec_())
