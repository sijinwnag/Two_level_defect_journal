# %%
import numpy as np
import sys, os
from PyQt5.QtWidgets import QApplication, QMainWindow, QAction, QActionGroup, QFileDialog, QGridLayout, QLineEdit, QLabel, QPushButton
from PyQt5 import QtGui, QtWidgets
from uiLifeSimu import Ui_LifeSimu
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
from semiconductor.recombination.intrinsic import Intrinsic
from semiconductor.recombination.extrinsic import SRH
from semiconductor.recombination.intrinsic import Radiative as rad
from semiconductor.recombination.intrinsic import Auger as aug
from semiconductor.material.intrinsic_carrier_density import IntrinsicCarrierDensity as NI
from semiconductor.electrical.ionisation import Ionisation as Ion
from semiconductor.material.thermal_velocity import ThermalVelocity as the_vel
from semiconductor.general_functions import carrierfunctions as CF
import scipy.constants as const
import openpyxl as xls

class LSimu(QMainWindow, Ui_LifeSimu):
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        self.setupUi(self)

        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.verticalLayout.addWidget(self.canvas)
        self.verticalLayout.addWidget(self.toolbar)
        self.ax1 = self.figure.add_subplot(111)
        self.ax1.set_xlabel(r'Excess carrier density $\rm [cm^{-3}]$')
        self.ax1.set_ylabel('Lifetime [s]')
        self.figure.tight_layout()

        self.lineEdit_Ndop.textChanged[str].connect(self.Ndopchange)
        self.lineEdit_CSRV.textChanged[str].connect(self.CSRVchange)
        self.lineEdit_Cj0.textChanged[str].connect(self.Cj0change)
        self.lineEdit_thickness.textChanged[str].connect(self.thicknesschange)
        self.lineEdit_Et_1.textChanged[str].connect(self.Et_1change)
        self.lineEdit_sign_1.textChanged[str].connect(self.sn_1change)
        self.lineEdit_sigp_1.textChanged[str].connect(self.sp_1change)
        self.lineEdit_Nt_1.textChanged[str].connect(self.Nt_1change)
        self.lineEdit_Et_2.textChanged[str].connect(self.Et_2change)
        self.lineEdit_sign_2.textChanged[str].connect(self.sn_2change)
        self.lineEdit_sigp_2.textChanged[str].connect(self.sp_2change)
        self.lineEdit_Nt_2.textChanged[str].connect(self.Nt_2change)
        self.lineEdit_temp.textChanged[str].connect(self.tempchange)
        self.lineEdit_nxc.textChanged[str].connect(self.nxcchange)
        self.comboBox_plottype.currentIndexChanged.connect(self.enableplotchoice)
        self.checkBox_dorid.stateChanged.connect(self.changesrhcheck)

        self.pushButton_plot.clicked.connect(self.plot)
        self.pushButton_exp.clicked.connect(self.export)

        self.menubar = self.menuBar()
        self.choosmodel = self.menubar.addMenu('Choose your models')
        self.nimodel = self.choosmodel.addMenu('ni models')
        self.Ionmodel = self.choosmodel.addMenu('Ionisation models')
        self.Radmodel = self.choosmodel.addMenu('Radiative models')
        self.Augmodel = self.choosmodel.addMenu('Auger models')
        self.themodel = self.choosmodel.addMenu('thermal velocity')

        self.nigroup = QActionGroup(self)
        self.radgroup = QActionGroup(self)
        self.auggroup = QActionGroup(self)
        self.iongroup = QActionGroup(self)
        self.thegroup = QActionGroup(self)

        self.thegroup.triggered.connect(self.updatetaus)

        availablNI = NI().available_models()
        availablRad = rad().available_models()
        availablAug = aug().available_models()
        availablIon = Ion().available_models()
        availablthervel = the_vel().available_models()

        self.SetRange = self.menubar.addMenu('Set calculation range')
        self.dnrangemenu = QAction('Injection range')
        self.SetRange.addAction(self.dnrangemenu)
        self.temprangemenu = QAction('Temperature range')
        self.SetRange.addAction(self.temprangemenu)
        self.dnrange = [12,17,50]
        self.temprange = [100,400,50]
        self.dnrangemenu.triggered.connect(self.changednrange)
        self.temprangemenu.triggered.connect(self.changetemprange)
        self.nxc = np.logspace(self.dnrange[0],self.dnrange[1],self.dnrange[2])
        self.Tlist = np.linspace(self.temprange[0],self.temprange[1],self.temprange[2])



        for nimodel in availablNI:
            a = self.nigroup.addAction(QAction(nimodel, checkable=True))
            if nimodel == 'Couderc_2014':
                a.setChecked(True)
            self.nimodel.addAction(a)
        for radmodel in availablRad:
            b = self.radgroup.addAction(QAction(radmodel, checkable=True))
            if radmodel == 'Altermatt_2005':
                b.setChecked(True)
            self.Radmodel.addAction(b)
        for augmodel in availablAug:
            c = self.auggroup.addAction(QAction(augmodel, checkable=True))
            if augmodel == 'Richter2012':
                c.setChecked(True)
            self.Augmodel.addAction(c)
        for ionmodel in availablIon:
            d = self.iongroup.addAction(QAction(ionmodel, checkable=True))
            if ionmodel == 'Altermatt_2006_table1':
                d.setChecked(True)
            self.Ionmodel.addAction(d)
        for themodel in availablthervel:
            a = self.thegroup.addAction(QAction(themodel, checkable=True))
            if themodel == 'Green_1990':
                a.setChecked(True)
            self.themodel.addAction(a)

        self.vth_e300, self.vth_h300 = the_vel().update(temp=300, author='Green_1990')
        self.vth_e300 = self.vth_e300[0]
        self.textinputdic = {'Ndop': 1e16, 'doptype':'n','CSRV': 10, 'Cj0': 10, 'thickness': 0.018,
                             'Et1': 0, 'sn1': 1e-13, 'sp1': 1e-13, 'Nt1': 1e12,
                             'Et2': 0.4, 'sn2': 1e-13, 'sp2': 1e-13, 'Nt2': 1e12,
                             'temp': 300, 'nxc': 1e12}
        self.lineEdit_taun_1.setText('{:.3e}'.format(1/self.vth_e300))
        self.lineEdit_taup_1.setText('{:.3e}'.format(1/self.vth_h300))
        self.lineEdit_taun_2.setText('{:.3e}'.format(1/self.vth_e300))
        self.lineEdit_taup_2.setText('{:.3e}'.format(1/self.vth_h300))
        self.idls_eff_tau=[]


    def changednrange(self):
        self.dialog = QtWidgets.QDialog()
        self.dialog.setWindowTitle('Set injection range for IDLS')
        grid = QGridLayout(self.dialog)
        self.dnstart = QLineEdit()
        self.dnend = QLineEdit()
        self.dn_n = QLineEdit()
        label1 = QLabel('From \t 1e')
        label2 = QLabel('To \t 1e')
        label3 = QLabel('with')
        label4 = QLabel('cm-3')
        label5 = QLabel('cm-3')
        label6 = QLabel('points')
        self.ok = QPushButton('OK')
        grid.addWidget(label1, 0, 0)
        grid.addWidget(self.dnstart, 0, 1)
        grid.addWidget(label4, 0, 2)
        grid.addWidget(label2, 1, 0)
        grid.addWidget(self.dnend, 1, 1)
        grid.addWidget(label5, 1, 2)
        grid.addWidget(label3, 2, 0)
        grid.addWidget(self.dn_n, 2, 1)
        grid.addWidget(label6, 2, 2)
        grid.addWidget(self.ok, 3, 1)
        self.dnstart.setText('{:.1f}'.format(self.dnrange[0]))
        self.dnend.setText('{:.1f}'.format(self.dnrange[1]))
        self.dn_n.setText('{:d}'.format(self.dnrange[-1]))

        self.dnstart.textChanged[str].connect(self.checkparam)
        self.dnend.textChanged[str].connect(self.checkparam)
        self.dn_n.textChanged[str].connect(self.checkparam)

        self.ok.clicked.connect(self.setparam)
        self.dialog.exec_()


    def checkparam(self):
        try:
            if float(self.dnstart.text()) > 0 and float(self.dnend.text()) > 0 and int(self.dn_n.text()) > 0:
                self.ok.setEnabled(True)
            else:
                self.ok.setEnabled(False)
        except ValueError:
            self.ok.setEnabled(False)

    def setparam(self):
        self.dnrange = [float(self.dnstart.text()),float(self.dnend.text()),int(self.dn_n.text())]
        self.nxc = np.logspace(self.dnrange[0],self.dnrange[1],self.dnrange[2])
        self.dialog.close()

    def changetemprange(self):
        self.dialog2 = QtWidgets.QDialog()
        self.dialog2.setWindowTitle('Set Temperature range for TDLS')
        grid = QGridLayout(self.dialog2)

        self.tempstart = QLineEdit()
        self.tempend = QLineEdit()
        self.temp_n = QLineEdit()
        label1 = QLabel('From ')
        label2 = QLabel('To ')
        label3 = QLabel('with')
        label4 = QLabel('K')
        label5 = QLabel('K')
        label6 = QLabel('points')
        self.ok2 = QPushButton('OK')
        grid.addWidget(label1, 0, 0)
        grid.addWidget(self.tempstart, 0, 1)
        grid.addWidget(label4, 0, 2)
        grid.addWidget(label2, 1, 0)
        grid.addWidget(self.tempend, 1, 1)
        grid.addWidget(label5, 1, 2)
        grid.addWidget(label3, 2, 0)
        grid.addWidget(self.temp_n, 2, 1)
        grid.addWidget(label6, 2, 2)
        grid.addWidget(self.ok2, 3, 1)
        self.tempstart.setText('{:.1f}'.format(self.temprange[0]))
        self.tempend.setText('{:.1f}'.format(self.temprange[1]))
        self.temp_n.setText('{:d}'.format(self.temprange[-1]))
        #
        self.tempstart.textChanged[str].connect(self.checkparam2)
        self.tempend.textChanged[str].connect(self.checkparam2)
        self.temp_n.textChanged[str].connect(self.checkparam2)
        #
        self.ok2.clicked.connect(self.setparam2)
        self.dialog2.exec_()


    def checkparam2(self):
        try:
            if float(self.tempstart.text()) > 0 and float(self.tempend.text()) > float(self.tempstart.text()) and int(self.temp_n.text()) > 0:
                self.ok2.setEnabled(True)
            else:
                self.ok2.setEnabled(False)
        except ValueError:
            self.ok2.setEnabled(False)

    def setparam2(self):
        self.temprange = [float(self.tempstart.text()),float(self.tempend.text()),int(self.temp_n.text())]
        self.Tlist = np.linspace(self.temprange[0],self.temprange[1],self.temprange[2])
        self.dialog2.close()


    def Ndopchange(self):
        try:
            lst = self.lineEdit_Ndop.text().split(',')
            lst = [float(i) for i in lst]
            self.textinputdic['Ndop'] = lst[0]
        except:
            self.textinputdic['Ndop'] = None

    def CSRVchange(self):
        try:
            lst = float(self.lineEdit_CSRV.text())
            self.textinputdic['CSRV'] = lst
        except:
            self.textinputdic['CSRV'] = None

    def Cj0change(self):
        try:
            lst = float(self.lineEdit_Cj0.text())
            self.textinputdic['Cj0'] = lst
        except:
            self.textinputdic['Cj0'] = None

    def thicknesschange(self):
        try:
            lst = float(self.lineEdit_thickness.text())
            self.textinputdic['thickness'] = lst
        except:
            self.textinputdic['thickness'] = None

    def Et_1change(self):
        try:
            lst = float(self.lineEdit_Et_1.text())
            self.textinputdic['Et1'] = lst
        except:
            self.textinputdic['Et1'] = None

    def sn_1change(self):
        try:
            lst = float(self.lineEdit_sign_1.text())
            self.textinputdic['sn1'] = lst
        except:
            self.textinputdic['sn1'] = None
        self.updatetaus()

    def sp_1change(self):
        try:
            lst = float(self.lineEdit_sigp_1.text())
            self.textinputdic['sp1'] = lst
        except:
            self.textinputdic['sp1'] = None
        self.updatetaus()

    def Nt_1change(self):
        try:
            lst = float(self.lineEdit_Nt_1.text())
            self.textinputdic['Nt1'] = lst
        except:
            self.textinputdic['Nt1'] = None
        self.updatetaus()

    def Et_2change(self):
        try:
            lst = float(self.lineEdit_Et_2.text())
            self.textinputdic['Et2'] = lst
        except:
            self.textinputdic['Et2'] = None

    def sn_2change(self):
        try:
            lst = float(self.lineEdit_sign_2.text())
            self.textinputdic['sn2'] = lst
        except:
            self.textinputdic['sn2'] = None
        self.updatetaus()

    def sp_2change(self):
        try:
            lst = float(self.lineEdit_sigp_2.text())
            self.textinputdic['sp2'] = lst
        except:
            self.textinputdic['sp2'] = None
        self.updatetaus()

    def Nt_2change(self):
        try:
            lst = float(self.lineEdit_Nt_2.text())
            self.textinputdic['Nt2'] = lst
        except:
            self.textinputdic['Nt2'] = None
        self.updatetaus()

    def tempchange(self):
        try:
            lst = float(self.lineEdit_temp.text())
            self.textinputdic['temp'] = lst
        except:
            self.textinputdic['temp'] = None
        self.updatetaus()

    def nxcchange(self):
        try:
            lst = float(self.lineEdit_nxc.text())
            self.textinputdic['nxc'] = lst
        except:
            self.textinputdic['nxc'] = None

    def updateallmodels(self):
        for action in self.nimodel.actions():
            if action.isChecked():
                self.ni_author = action.text()
        for action in self.Ionmodel.actions():
            if action.isChecked():
                self.ionauthor = action.text()
        for action in self.themodel.actions():
            if action.isChecked():
                self.vth_author = action.text()
        for action in self.Radmodel.actions():
            if action.isChecked():
                self.radauthor = action.text()
        for action in self.Augmodel.actions():
            if action.isChecked():
                self.augauthor = action.text()

    def changesrhcheck(self):
        self.updatetaus()
        if self.checkBox_dorid.isChecked()==True:
            self.checkBox_1stDefect.setChecked(True)
            self.checkBox_2ndDefect.setChecked(True)
            self.checkBox_1stDefect.setEnabled(False)
            self.checkBox_2ndDefect.setEnabled(False)
            self.label_31.setVisible(False)
            self.lineEdit_Nt_2.setVisible(False)
            self.label_26.setVisible(False)
        else:
            self.label_31.setVisible(True)
            self.lineEdit_Nt_2.setVisible(True)
            self.label_26.setVisible(True)
            self.checkBox_1stDefect.setEnabled(True)
            self.checkBox_2ndDefect.setEnabled(True)

    def updatetaus(self):
        self.calc_basics_idls()
        if self.textinputdic['temp'] is not None and self.textinputdic['Nt1'] is not None and self.textinputdic['sn1'] is not None:
            taun1 = 1/self.vth_e/self.textinputdic['Nt1']/self.textinputdic['sn1']
            self.lineEdit_taun_1.setText('{:.3e}'.format(taun1))
        else:
            self.lineEdit_taun_1.clear()
        if self.textinputdic['temp'] is not None and self.textinputdic['Nt1'] is not None and self.textinputdic['sp1'] is not None:
            taup1 = 1/self.vth_e/self.textinputdic['Nt1']/self.textinputdic['sp1']
            self.lineEdit_taup_1.setText('{:.3e}'.format(taup1))
        else:
            self.lineEdit_taup_1.clear()
        if self.checkBox_dorid.isChecked()==True:
            if self.textinputdic['temp'] is not None and self.textinputdic['Nt1'] is not None and self.textinputdic['sn2'] is not None:
                taun2 = 1/self.vth_e/self.textinputdic['Nt1']/self.textinputdic['sn2']
                self.lineEdit_taun_2.setText('{:.3e}'.format(taun2))
            else:
                self.lineEdit_taun_2.clear()
            if self.textinputdic['temp'] is not None and self.textinputdic['Nt1'] is not None and self.textinputdic['sp2'] is not None:
                taup2 = 1/self.vth_e/self.textinputdic['Nt1']/self.textinputdic['sp2']
                self.lineEdit_taup_2.setText('{:.3e}'.format(taup2))
            else:
                self.lineEdit_taup_2.clear()
        else:
            if self.textinputdic['temp'] is not None and self.textinputdic['Nt2'] is not None and self.textinputdic['sn2'] is not None:
                taun2 = 1/self.vth_e/self.textinputdic['Nt2']/self.textinputdic['sn2']
                self.lineEdit_taun_2.setText('{:.3e}'.format(taun2))
            else:
                self.lineEdit_taun_2.clear()
            if self.textinputdic['temp'] is not None and self.textinputdic['Nt2'] is not None and self.textinputdic['sp2'] is not None:
                taup2 = 1/self.vth_e/self.textinputdic['Nt2']/self.textinputdic['sp2']
                self.lineEdit_taup_2.setText('{:.3e}'.format(taup2))
            else:
                self.lineEdit_taup_2.clear()

    def calc_basics_tdls(self):
        self.updateallmodels()
        self.vth_elist=[]
        self.vth_hlist=[]
        self.Nalist=[]
        self.Ndlist=[]
        self.nilist=[]
        for T in self.Tlist:
            ni = NI().update(temp=T, author=self.ni_author)
            ni = ni[0]
            vth_e, vth_h = the_vel().update(temp=T, author=self.vth_author)
            try:
                vth_e = vth_e[0]
            except:
                pass
            self.nilist.append(ni)
            self.vth_elist.append(vth_e)
            self.vth_hlist.append(vth_h)
        if self.textinputdic['nxc'] is not None and self.textinputdic['Ndop'] is not None:
            Ndop = self.textinputdic['Ndop']
            if self.radioButton_n.isChecked() == True:
                for T in self.Tlist:
                    self.Ndlist.append(Ion(temp=T, ni_author=self.ni_author).update_dopant_ionisation(
                        author=self.ionauthor, N_dop=Ndop, nxc=self.textinputdic['nxc'], impurity='phosphorous'))
                    self.Nalist.append(1)
            if self.radioButton_p.isChecked() == True:
                for T in self.Tlist:
                    self.Nalist.append(Ion(temp=T, ni_author=self.ni_author).update_dopant_ionisation(
                        author=self.ionauthor, N_dop=Ndop, nxc=self.textinputdic['nxc'], impurity='boron'))
                    self.Ndlist.append(1)
        else:
            self.Nalist=[]
            self.Ndlist=[]

    def calc_tdls_intrinsic(self):
        self.tdls_intrinsic_itau=[]
        self.tdls_intrinsic_tau=[]

    def calc_tdls_srf(self):
        self.tdls_srf_itau = []
        self.tdls_srf_tau = []

    def calc_tdls_srh(self):
        self.tdls_srh_itau1 = []
        self.tdls_srh_itau2 = []
        self.tdls_srh_tau1 = []
        self.tdls_srh_tau2 = []
        self.tdls_srh_itau12 = []
        self.tdls_srh_tau12 = []
        self.calc_basics_tdls()
        if self.checkBox_dorid.isChecked() == True:
            self.tdls_srh_itau1 = []
            self.tdls_srh_itau2 = []
            self.tdls_srh_tau1 = []
            self.tdls_srh_tau2 = []
            if self.textinputdic['nxc'] is not None and len(self.Nalist)>0 and len(self.nilist)>0 and\
                len(self.vth_elist)>0 and len(self.vth_hlist)>0 and \
                self.textinputdic['Et1'] is not None and self.textinputdic['Et2'] is not None and \
                self.textinputdic['Nt1'] is not None and  \
                self.textinputdic['sn1'] is not None and self.textinputdic['sn2'] is not None and \
                self.textinputdic['sp1'] is not None and self.textinputdic['sp2'] is not None:
                Et1 = self.textinputdic['Et1']
                Et2 = self.textinputdic['Et2']
                Nt = self.textinputdic['Nt1']
                alpha_e1 = self.textinputdic['sn1']*np.asarray(self.vth_elist)
                alpha_h1 = self.textinputdic['sp1']*np.asarray(self.vth_hlist)
                alpha_e2 = self.textinputdic['sn2']*np.asarray(self.vth_elist)
                alpha_h2 = self.textinputdic['sp2']*np.asarray(self.vth_hlist)
                for T, ni, Na, Nd, ae1, ah1,ae2,ah2 in zip(self.Tlist,self.nilist,self.Nalist,self.Ndlist,alpha_e1,alpha_h1,alpha_e2,alpha_h2):
                    self.tdls_srh_itau12.append(self.twolevelSRH_full(self.textinputdic['nxc'],T,ni,Na,Nd,Nt,Et1,Et2,ae1, ah1, ae2, ah2))
                self.tdls_srh_itau12=np.asarray(self.tdls_srh_itau12)
                self.tdls_srh_tau12 = 1/self.tdls_srh_itau12
            else:
                self.tdls_srh_itau12 = []
                self.tdls_srh_tau12 = []
        else:
            self.tdls_srh_itau12 = []
            self.tdls_srh_tau12 = []
            if self.checkBox_1stDefect.isChecked()==True:
                if self.textinputdic['nxc'] is not None and len(self.Nalist)>0 and\
                    len(self.vth_elist)>0 and len(self.vth_hlist)>0 and len(self.nilist)>0 and\
                    self.textinputdic['Et1'] is not None and self.textinputdic['Nt1'] is not None and \
                    self.textinputdic['sn1'] is not None and self.textinputdic['sp1'] is not None:
                    Et = self.textinputdic['Et1']
                    tau_e = 1./self.textinputdic['Nt1']/np.asarray(self.vth_elist)/self.textinputdic['sn1']
                    tau_h = 1./self.textinputdic['Nt1']/np.asarray(self.vth_hlist)/self.textinputdic['sp1']
                    for ni, te, th, T, Na, Nd in zip(self.nilist,tau_e,tau_h,self.Tlist,self.Nalist,self.Ndlist):
                        self.tdls_srh_itau1.append(self.SRHtau(self.textinputdic['nxc'], ni, Et, te, th, T, Na, Nd))
                    self.tdls_srh_itau1 = np.asarray(self.tdls_srh_itau1)
                    self.tdls_srh_tau1=1/self.tdls_srh_itau1
                else:
                    self.tdls_srh_itau1=[]
                    self.tdls_srh_tau1=[]
            else:
                self.tdls_srh_itau1=[]
                self.tdls_srh_tau1=[]
            if self.checkBox_2ndDefect.isChecked()==True:
                if self.textinputdic['nxc'] is not None and len(self.Nalist)>0 and\
                    len(self.vth_elist)>0 and len(self.vth_hlist)>0 and len(self.nilist)>0 and\
                    self.textinputdic['Et2'] is not None and self.textinputdic['Nt2'] is not None and \
                    self.textinputdic['sn2'] is not None and self.textinputdic['sp2'] is not None:
                    Et = self.textinputdic['Et2']
                    tau_e = 1./self.textinputdic['Nt2']/np.asarray(self.vth_elist)/self.textinputdic['sn2']
                    tau_h = 1./self.textinputdic['Nt2']/np.asarray(self.vth_hlist)/self.textinputdic['sp2']
                    for ni, te, th, T, Na, Nd in zip(self.nilist,tau_e,tau_h,self.Tlist,self.Nalist,self.Ndlist):
                        self.tdls_srh_itau2.append(self.SRHtau(self.textinputdic['nxc'], ni, Et, te, th, T, Na, Nd))
                    self.tdls_srh_itau2 = np.asarray(self.tdls_srh_itau2)
                    self.tdls_srh_tau2=1/self.tdls_srh_itau2
                else:
                    self.tdls_srh_itau2=[]
                    self.tdls_srh_tau2=[]
            else:
                self.tdls_srh_itau2=[]
                self.tdls_srh_tau2=[]


    def calc_tdls_effbyT(self):
        self.calc_tdls_intrinsic()
        self.calc_tdls_srf()
        self.calc_tdls_srh()
        a=0*self.Tlist
        if len(self.tdls_srh_itau1)>0:
            self.tdls_srh_itau1 =self.tdls_srh_itau1.reshape(self.Tlist.shape)*self.Tlist
            self.tdls_srh_tau1  = 1/self.tdls_srh_itau1
            a+=self.tdls_srh_itau1
        if len(self.tdls_srh_itau2)>0:
            self.tdls_srh_itau2 = self.tdls_srh_itau2.reshape(self.Tlist.shape)*self.Tlist
            self.tdls_srh_tau2  = 1/self.tdls_srh_itau2
            a+=self.tdls_srh_itau2
        if len(self.tdls_srh_itau12)>0:
            self.tdls_srh_itau12 = self.tdls_srh_itau12.reshape(self.Tlist.shape)*self.Tlist
            self.tdls_srh_tau12  = 1/self.tdls_srh_itau12
            a+=self.tdls_srh_itau12
        if np.count_nonzero(a)>0:
            self.tdls_eff_itau = a
            self.tdls_eff_tau = 1/self.tdls_eff_itau
        else:
            self.tdls_eff_itau = []
            self.tdls_eff_tau =[]
        if len(self.tdls_srh_itau1)>0 and len(self.tdls_srh_itau2)>0:
            self.tdls_srh_itau = self.tdls_srh_itau1 +self.tdls_srh_itau2
            self.tdls_srh_tau = 1/self.tdls_srh_itau
        else:
            self.tdls_srh_itau =[]
            self.tdls_srh_tau =[]

    def calc_basics_idls(self):
        self.updateallmodels()
        if self.textinputdic['temp'] is not None:
            temp = self.textinputdic['temp']
            self.ni = NI().update(temp=temp, author=self.ni_author)
            self.ni = self.ni[0]
            self.vth_e, self.vth_h = the_vel().update(temp=temp, author=self.vth_author)
            try:
                self.vth_e = self.vth_e[0]
            except:
                pass
            self.vth_e300, self.vth_h300 = the_vel().update(temp=300, author=self.vth_author)
            try:
                self.vth_e300 = self.vth_e300[0]
            except:
                pass
        else:
            self.vth_e = None
            self.vth_h = None
            self.ni = None

        if self.textinputdic['temp'] is not None and self.textinputdic['Ndop'] is not None:
            Ndop = self.textinputdic['Ndop']
            temp = self.textinputdic['temp']
            if self.radioButton_n.isChecked() == True:
                self.Nd = Ion(temp=temp, ni_author=self.ni_author).update_dopant_ionisation(
                    author=self.ionauthor, N_dop=Ndop, nxc=self.nxc, impurity='phosphorous')
                self.Na = 0*self.nxc+1
            if self.radioButton_p.isChecked() == True:
                self.Na = Ion(temp=temp, ni_author=self.ni_author).update_dopant_ionisation(
                    author=self.ionauthor, N_dop=Ndop, nxc=self.nxc, impurity='boron')
                self.Nd = 0*self.nxc+1
        else:
            self.Na = []
            self.Nd = []

    def calc_idls_intrinsic(self):
        if not self.radioButton_nointrin.isChecked():
            self.calc_basics_idls()
            if len(self.Na)>0 and len(self.Nd)>0 and self.textinputdic['temp'] is not None:
                temp = self.textinputdic['temp']
                self.idls_rad_itau = rad().itau(ni_author=self.ni_author,
                                                      author=self.radauthor,
                                                      temp=temp,
                                                      Na=self.Na,
                                                      Nd=self.Nd,
                                                      nxc=self.nxc)
                self.idls_aug_itau = aug().itau(ni_author=self.ni_author,
                                                       author=self.augauthor,
                                                       temp=temp,
                                                       Na=self.Na,
                                                       Nd=self.Nd,
                                                       nxc=self.nxc)
                self.idls_intrinsic_itau = rad().itau(ni_author=self.ni_author,
                                                      author=self.radauthor,
                                                      temp=temp,
                                                      Na=self.Na,
                                                      Nd=self.Nd,
                                                      nxc=self.nxc) + \
                                            aug().itau(ni_author=self.ni_author,
                                                       author=self.augauthor,
                                                       temp=temp,
                                                       Na=self.Na,
                                                       Nd=self.Nd,
                                                       nxc=self.nxc)
                self.idls_intrinsic_tau = 1/self.idls_intrinsic_itau
                self.idls_rad_tau = 1/self.idls_rad_itau
                self.idls_aug_tau = 1/self.idls_aug_itau
            else:
                self.idls_intrinsic_itau = []
                self.idls_intrinsic_tau = []
                self.idls_rad_itau = []
                self.idls_rad_tau = []
                self.idls_aug_itau = []
                self.idls_aug_tau = []
        else:
            self.idls_intrinsic_itau = []
            self.idls_intrinsic_tau = []
            self.idls_rad_itau = []
            self.idls_rad_tau = []
            self.idls_aug_itau = []
            self.idls_aug_tau = []

    def calc_idls_srf(self):
        if self.radioButton_nosrv.isChecked() == True:
            self.idls_srf_itau = []
            self.idls_srf_tau = []
        if self.radioButton_CSRV.isChecked() == True:
            self.calc_basics_idls()
            if self.textinputdic['thickness'] is not None and self.textinputdic['CSRV'] is not None:
                self.idls_srf_itau = self.textinputdic['CSRV']/self.textinputdic['thickness']+0*self.nxc
                self.idls_srf_tau = 1/self.idls_srf_itau
            else:
                self.idls_srf_itau = []
                self.idls_srf_tau = []
        if self.radioButton_CJ0.isChecked() == True:
            self.calc_basics_idls()
            if self.textinputdic['thickness'] is not None and self.textinputdic['Cj0'] is not None and self.ni and len(self.Na)>0:
                self.idls_srf_itau = self.textinputdic['Cj0']*1e-15*(self.Na+self.Nd+self.nxc)/self.textinputdic['thickness']/const.e/self.ni**2
                self.idls_srf_tau = 1/self.idls_srf_itau
            else:
                self.idls_srf_itau = []
                self.idls_srf_tau = []

    def SRHtau(self, nxc, ni, Et, tau_e, tau_h, T, Na, Nd, **kwargs):
        nh1 = ni * np.exp(-Et * const.e / (const.k * T))
        ne1 = ni * np.exp(Et * const.e / (const.k * T))
        ne, nh = CF.get_carriers(Na=Na, Nd=Nd, nxc=nxc, temp=T, ni=ni)
        U = (ne * nh - ni**2) / (tau_h * (ne + ne1) + tau_e * (nh + nh1))
        return U/nxc

    def twolevelSRH_full(self,nxc, T, ni, Na, Nd, Nt, Et1, Et2, alpha_e1, alpha_h1, alpha_e2, alpha_h2, **kwarg):
        n, p = CF.get_carriers(Na=Na, Nd=Nd, nxc=nxc, temp=T, ni=ni)
        n1 = ni * np.exp(Et1* const.e / const.k / T)
        p1 = ni * np.exp(-Et1 * const.e/ const.k / T)
        n2 = ni * np.exp(Et2 * const.e/ const.k / T)
        p2 = ni * np.exp(-Et2 * const.e/ const.k / T)
        R = Nt * (n * p - ni**2) / \
            (1 + ((alpha_e1 * n1 + alpha_h1 * p) / (alpha_e1 * n + alpha_h1 * p1)) +\
            ((alpha_e2 * n + alpha_h2 * p2) / (alpha_e2 *n2 + alpha_h2 * p))) *\
            ((alpha_e1 * alpha_h1 / (alpha_e1 * n + alpha_h1 * p1)) +\
            (alpha_e2 * alpha_h2 / (alpha_e2 * n2 + alpha_h2 * p)))

        itau = R/nxc
        return itau

    def calc_idls_srh(self):
        self.calc_basics_idls()
        if self.checkBox_dorid.isChecked() == True:
            self.idls_srh_itau1 = []
            self.idls_srh_itau2 = []
            self.idls_srh_tau1 = []
            self.idls_srh_tau2 = []
            self.idls_srh_tau = []
            self.idls_srh_tau = []
            if self.textinputdic['temp'] is not None and len(self.Na)>0 and self.ni is not None and\
                self.vth_e is not None and self.vth_h is not None and \
                self.textinputdic['Et1'] is not None and self.textinputdic['Et2'] is not None and \
                self.textinputdic['Nt1'] is not None and  \
                self.textinputdic['sn1'] is not None and self.textinputdic['sn2'] is not None and \
                self.textinputdic['sp1'] is not None and self.textinputdic['sp2'] is not None:
                Et1 = self.textinputdic['Et1']
                Et2 = self.textinputdic['Et2']
                Nt = self.textinputdic['Nt1']
                alpha_e1 = self.textinputdic['sn1']*self.vth_e
                alpha_h1 = self.textinputdic['sp1']*self.vth_h
                alpha_e2 = self.textinputdic['sn2']*self.vth_e
                alpha_h2 = self.textinputdic['sp2']*self.vth_h
                self.idls_srh_itau12 =self.twolevelSRH_full(self.nxc,self.textinputdic['temp'],self.ni,self.Na,self.Nd,Nt,Et1,Et2,alpha_e1, alpha_h1, alpha_e2, alpha_h2)
                self.idls_srh_tau12 = 1/self.idls_srh_itau12
            else:
                self.idls_srh_itau12 = []
                self.idls_srh_tau12 = []
        else:
            self.idls_srh_itau12 = []
            self.idls_srh_tau12 = []
            if self.checkBox_1stDefect.isChecked()==True:
                if self.textinputdic['temp'] is not None and len(self.Na) and\
                    self.vth_e is not None and self.vth_h is not None and \
                    self.textinputdic['Et1'] is not None and self.textinputdic['Nt1'] is not None and \
                    self.textinputdic['sn1'] is not None and self.textinputdic['sp1'] is not None:
                    Et = self.textinputdic['Et1']
                    tau_e = 1./self.textinputdic['Nt1']/self.vth_e/self.textinputdic['sn1']
                    tau_h = 1./self.textinputdic['Nt1']/self.vth_h/self.textinputdic['sp1']
                    self.idls_srh_itau1=self.SRHtau(self.nxc, self.ni, Et, tau_e, tau_h, self.textinputdic['temp'], self.Na, self.Nd)
                    self.idls_srh_tau1=1/self.idls_srh_itau1
                else:
                    self.idls_srh_itau1=[]
                    self.idls_srh_tau1=[]
            else:
                self.idls_srh_itau1=[]
                self.idls_srh_tau1=[]
            if self.checkBox_2ndDefect.isChecked()==True:
                if self.textinputdic['temp'] is not None and len(self.Na)>0 and\
                    self.vth_e is not None and self.vth_h is not None and \
                    self.textinputdic['Et2'] is not None and self.textinputdic['Nt2'] is not None and \
                    self.textinputdic['sn2'] is not None and self.textinputdic['sp2'] is not None:
                    Et = self.textinputdic['Et2']
                    tau_e = 1./self.textinputdic['Nt2']/self.vth_e/self.textinputdic['sn2']
                    tau_h = 1./self.textinputdic['Nt2']/self.vth_h/self.textinputdic['sp2']
                    self.idls_srh_itau2=self.SRHtau(self.nxc, self.ni, Et, tau_e, tau_h, self.textinputdic['temp'], self.Na, self.Nd)
                    self.idls_srh_tau2=1/self.idls_srh_itau2
                else:
                    self.idls_srh_itau2=[]
                    self.idls_srh_tau2=[]
            else:
                self.idls_srh_itau2=[]
                self.idls_srh_tau2=[]
            if len(self.idls_srh_itau1)>0 and len(self.idls_srh_itau2)>0:
                self.idls_srh_itau = self.idls_srh_itau1 +self.idls_srh_itau2
                self.idls_srh_tau = 1/self.idls_srh_itau
            else:
                self.idls_srh_itau =[]
                self.idls_srh_tau =[]


    def calc_idls_eff(self):
        self.calc_idls_intrinsic()
        self.calc_idls_srf()
        self.calc_idls_srh()
        a=0*self.nxc
        # print(np.shape(a))
        # print(np.shape(self.idls_srh_itau1))
        if len(self.idls_srh_itau1)>0:
            a+=self.idls_srh_itau1
        if len(self.idls_srh_itau2)>0:
            a+=self.idls_srh_itau2
        if len(self.idls_srh_itau12)>0:
            a+=self.idls_srh_itau12
        if len(self.idls_srf_itau)>0:
            a+=self.idls_srf_itau
        if len(self.idls_intrinsic_itau)>0:
            a+=self.idls_intrinsic_itau
        if np.count_nonzero(a)>0:
            self.idls_eff_itau = a
            self.idls_eff_tau = 1/self.idls_eff_itau
        else:
            self.idls_eff_itau = []
            self.idls_eff_tau =[]

    def calc_XY(self):
        self.calc_basics_idls()
        if len(self.Na)>0 and self.textinputdic['temp'] is not None:
            ne, nh = CF.get_carriers(Na=self.Na, Nd=self.Nd, nxc=self.nxc, temp=self.textinputdic['temp'], ni=self.ni)
            if self.radioButton_n.isChecked() == True:
                self.X=nh/ne
            elif self.radioButton_p.isChecked() == True:
                self.X=ne/nh
        else:
            self.X=None

    def enableplotchoice(self):
        if self.comboBox_plottype.currentText()=='IDLS':
            self.comboBox_plotopt.setEnabled(True)
            self.LSimu_SRV_Group.setEnabled(True)
            self.groupBox.setEnabled(True)
        else:
            self.comboBox_plotopt.setEnabled(False)
            self.radioButton_nosrv.setChecked(True)
            self.LSimu_SRV_Group.setEnabled(False)
            self.radioButton_nointrin.setChecked(True)
            self.groupBox.setEnabled(False)

    def plot(self):
        if self.comboBox_plottype.currentText()=='IDLS':
            self.plotidls()
        else:
            self.plottdls()

    def plotidls(self):
        self.calc_idls_eff()
        self.ax1.clear()
        self.ax1.grid()
        if self.comboBox_plotopt.currentText()=='Lifetime vs nxc':
            self.ax1.set_xlabel(r'Excess carrier density $\rm [cm^{-3}]$')
            self.ax1.set_ylabel(r'Lifetime [s]')
            self.plotxy(self.nxc,self.idls_eff_tau,self.idls_rad_tau,self.idls_aug_tau,self.idls_srf_tau,self.idls_srh_tau1,self.idls_srh_tau2,self.idls_srh_tau12,self.idls_srh_tau)
            self.ax1.loglog()
        if self.comboBox_plotopt.currentText()=='Inverse lifetime vs nxc':
            self.ax1.set_xlabel(r'Excess carrier density $\rm [cm^{-3}]$')
            self.ax1.set_ylabel(r'Inverse Lifetime $\rm [s^{-1}]$')
            self.plotxy(self.nxc,self.idls_eff_itau,self.idls_intrinsic_itau,self.idls_srf_itau,self.idls_srh_itau1,self.idls_srh_itau2,self.idls_srh_itau12,self.idls_srh_itau)
            self.ax1.semilogx()
        if self.comboBox_plotopt.currentText()=='Lifetime vs X(Y)':
            self.calc_XY()
            self.ax1.set_xlabel(r'X/Y')
            self.ax1.set_ylabel(r'Lifetime [s]')
            self.plotxy(self.X,self.idls_eff_tau,self.idls_intrinsic_tau,self.idls_srf_tau,self.idls_srh_tau1,self.idls_srh_tau2,self.idls_srh_tau12,self.idls_srh_tau)
        if self.comboBox_plotopt.currentText()=='Inverse lifetime vs X(Y)':
            self.calc_XY()
            self.ax1.set_xlabel(r'X/Y')
            self.ax1.set_ylabel(r'Inverse Lifetime $\rm [s^{-1}]$')
            self.plotxy(self.X,self.idls_eff_itau,self.idls_intrinsic_itau,self.idls_srf_itau,self.idls_srh_itau1,self.idls_srh_itau2,self.idls_srh_itau12,self.idls_srh_itau)

        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()

    def plottdls(self):
        self.calc_tdls_effbyT()
        self.ax1.clear()
        self.ax1.grid()
        self.ax1.set_xlabel(r'1000/T $\rm [K^{-1}]$')
        self.ax1.set_ylabel(r'$\tau/T$ [s/K]')
        self.plotxy(1000/np.asarray(self.Tlist),np.asarray(self.tdls_eff_tau),\
                    np.asarray(self.tdls_intrinsic_tau),np.asarray(self.tdls_srf_tau),\
                    np.asarray(self.tdls_srh_tau1),np.asarray(self.tdls_srh_tau2),\
                    np.asarray(self.tdls_srh_tau12),np.asarray(self.tdls_srh_tau))
        self.ax1.semilogy()
        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()

    def plotxy(self,x,eff,rad,aug,srf,srh1,srh2,srh12,srh):
        if len(eff)>0 and self.checkBox_eff.isChecked():
            self.ax1.plot(x,eff,'k',label='Effective')
        if len(rad)>0 and len(aug)>0 and self.checkBox_int.isChecked():
            self.ax1.plot(x,rad,'g-.',label='Radiative')
            self.ax1.plot(x,aug,'r-.',label='Auger')
        if len(srf)>0 and self.checkBox_SRV.isChecked():
            self.ax1.plot(x,srf,'b-.',label='Surface')
        if self.checkBox_SRH.isChecked():
            if len(srh12)>0:
                self.ax1.plot(x,srh12,'r-.',label='Two level defect')
            if len(srh1)>0:
                self.ax1.plot(x,srh1,'m--',label='Defect 1')
            if len(srh2)>0:
                self.ax1.plot(x,srh2,'c--',label='Defect 2')
            if len(srh)>0:
                self.ax1.plot(x,srh,'r-.',label='Two defects')

    def export(self):
        fname = QFileDialog.getSaveFileName(parent=self,caption='save data',filter="save plotted data to csv file (*.csv);;save effective idls to fake sinton file (*.xlsm)")
        if fname[-1] == "save plotted data to csv file (*.csv)":
            self.expcsv(fname[0])
        elif fname[-1] == "save effective idls to fake sinton file (*.xlsm)":
            self.expxlsm(fname[0])


    def expcsv(self,fname):
        # print(self.ax1.xaxis.get_label().get_text())
        if self.ax1.yaxis.get_label().get_text()=='Lifetime [s]':
            yhi='tau_'
        elif self.ax1.yaxis.get_label().get_text()==r'Inverse Lifetime $\rm [s^{-1}]$':
            yhi='itau_'
        elif self.ax1.yaxis.get_label().get_text()==r'$\tau/T$ [s/K]':
            yhi='tau/T_'

        if self.ax1.xaxis.get_label().get_text()==r'Excess carrier density $\rm [cm^{-3}]$':
            d = [self.nxc]
            header = 'nxc'
        elif self.ax1.xaxis.get_label().get_text()==r'X/Y':
            d = [self.X]
            header = 'X/Y'
        elif self.ax1.xaxis.get_label().get_text()==r'1000/T $\rm [K^{-1}]$':
            d = [self.Tlist]
            header = 'Temp'

        lines = self.ax1.lines
        for l in lines:
            y =l.get_ydata()
            header += ','+yhi+l.get_label()
            d.append(y)
        np.savetxt(fname,np.asarray(d).T,delimiter=',',header=header)

    def expxlsm(self,fname):
        if self.textinputdic['temp'] is not None and self.textinputdic['Ndop'] is not None and len(self.idls_eff_tau)>0:
            p, name = os.path.split(fname)
            if self.radioButton_n.isChecked() == True:
                doptype='n'
            if self.radioButton_p.isChecked() == True:
                doptype='p'
            name=name[:-5]
            book = xls.load_workbook(os.path.dirname(os.path.realpath(__file__))+"./dummy.xlsm", keep_vba=True)
            sheet1 = book.get_sheet_by_name("RawData")
            sheet2 = book.get_sheet_by_name("User")
            sheet2['A6'].value = name
            sheet2['J9'].value = self.textinputdic['Ndop']
            sheet2['D6'].value = doptype
            if self.textinputdic['thickness'] is not None:
                sheet2['B6'].value = self.textinputdic['thickness']
            sheet2['L9'].value = self.textinputdic['temp'] - 273.15
            for i in range(len(self.nxc)):
                sheet1.cell(row=5 + i, column=7).value = self.nxc[i]
                sheet1.cell(row=5 + i, column=5).value = self.idls_eff_tau[i]
            book.save(fname)




if __name__ == '__main__':
    app = QApplication(sys.argv)
    LSimu = LSimu()
    LSimu.show()
    sys.exit(app.exec_())
