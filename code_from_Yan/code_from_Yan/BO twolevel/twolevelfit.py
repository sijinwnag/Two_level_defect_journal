# -*- coding: utf-8 -*-
"""
Created on Thu Apr 26 14:40:38 2018

@author: z5093358
"""
# %%
import numpy as np
import matplotlib.pyplot as plt
import scipy.constants as const
import matplotlib
from semiconductor.material.thermal_velocity import ThermalVelocity as Vel_th
from semiconductor.material.intrinsic_carrier_density import IntrinsicCarrierDensity as NI
#from semiconductor.electrical.mobility import Mobility as mob
from semiconductor.electrical.ionisation import Ionisation as Ion
from semiconductor.general_functions import carrierfunctions as CF
from scipy.optimize import minimize
import matplotlib.colors as colors
import openpyxl as xls

matplotlib.pyplot.switch_backend('Qt5Agg')
kb = const.k / const.e

ve300, vh300 = Vel_th().update(temp=300)
ve300 = ve300[0]

def twolevelSRH_full(nxc, Ndop, T, doptype, Nt, Et1, Et2, sigma_e1, sigma_e2, sigma_h1, sigma_h2, **kwarg):
    ni = NI().update(temp=T)
    ni = ni[0]
    ve, vh = Vel_th().update(temp=T)
    ve = ve[0]
    if doptype == 'p':
        Na = Ion(temp=T).update_dopant_ionisation(
            N_dop=Ndop, nxc=0, impurity='boron')
        Na = Na[0]
        Nd = 0
    elif doptype == 'n':
        Nd = Ion(temp=T).update_dopant_ionisation(
            N_dop=Ndop, nxc=0, impurity='phosphorous')
        Nd = Nd[0]
        Na = 0
    n0, p0 = CF.get_carriers(Na=Na, Nd=Nd, nxc=0, temp=T, ni=ni)
    alpha_e1 = sigma_e1 * ve
    alpha_h1 = sigma_h1 * vh
    alpha_e2 = sigma_e2 * ve
    alpha_h2 = sigma_h2 * vh
    n1 = ni * np.exp(Et1 / kb / T)
    p1 = ni * np.exp(-Et1 / kb / T)
    n2 = ni * np.exp(Et2 / kb / T)
    p2 = ni * np.exp(-Et2 / kb / T)
    n = n0 + nxc
    p = p0 + nxc
    R = Nt * (n * p - ni**2) / (1 + ((alpha_e1 * n1 + alpha_h1 * p) / (alpha_e1 * n + alpha_h1 * p1)) + ((alpha_e2 * n + alpha_h2 * p2) / (alpha_e2 *
                                                                                                                                       n2 + alpha_h2 * p))) * ((alpha_e1 * alpha_h1 / (alpha_e1 * n + alpha_h1 * p1)) + (alpha_e2 * alpha_h2 / (alpha_e2 * n2 + alpha_h2 * p)))
    tau = nxc / R
    return tau


filelist = ['n30C.xlsm', 'p30C.xlsm']
Doplist = []
Tlist = []
Typelist=[]
nxclist=[]
taulist=[]

for fname in filelist:
    nxc=[]
    tau=[]
    wb = xls.load_workbook(filename=fname, data_only=True)
    ws = wb['RawData']
    lenth = ws.max_row
    for row in ws.iter_rows('E5:G' + str(lenth)):
        if row[0].value is not None and row[-1].value is not None:
            tau.append(row[0].value)
            nxc.append(row[-1].value)
    nxc = np.asarray(nxc)
    tau = np.asarray(tau)
    idx = np.argsort(nxc)
    nxc = nxc[idx]
    tau = tau[idx]
    ws1 = wb['User']
    temp = ws1['L9'].value + 273.15
    Tlist.append(temp)
    temp = ws1['J9'].value
    Doplist.append(temp)
    temp = ws1['D6'].value
    Typelist.append(temp)
    nxclist.append(nxc)
    taulist.append(tau)
    
nlist=[]
plist=[]
nilist=[]
velist=[]
vhlist=[]

for doptype,nxc,T,Ndop in zip(Typelist,nxclist,Tlist,Doplist):
    ni = NI().update(temp=T)
    ni = ni[0]
    ve, vh = Vel_th().update(temp=T)
    ve = ve[0]
    nilist.append(ni)
    velist.append(ve)
    vhlist.append(vh)
    if doptype == 'p':
        Na = Ion(temp=T).update_dopant_ionisation(
            N_dop=Ndop, nxc=0, impurity='boron')
        Na = Na[0]
        Nd = 0
    elif doptype == 'n':
        Nd = Ion(temp=T).update_dopant_ionisation(
            N_dop=Ndop, nxc=0, impurity='phosphorous')
        Nd = Nd[0]
        Na = 0
    n0, p0 = CF.get_carriers(Na=Na, Nd=Nd, nxc=0, temp=T, ni=ni)
    n0=n0[0]
    p0=p0[0]
    n = n0 + nxc
    p = p0 + nxc
    nlist.append(n)
    plist.append(p)


Et1list=np.linspace(-0.6,0.6,50)
Et2list=np.linspace(-0.6,0.6,50)
#Et1list=[0]
#Et2list=[0]

taun1list=[]
taun2list=[]
k1list=[]
k2list=[]
scalelist=[]
residuallist=[]
Etlist=[]

Nt=1e12 
count=0
for Et1 in Et1list:
    for Et2 in Et2list:
        Etlist.append([Et1,Et2])
        def squareresidual(siglist, **kwarg):
            taun1=siglist[0]*1e-5
            taun2=siglist[1]*1e-5
            k1=siglist[2]
            k2=siglist[3]
            fraction = siglist[4]
            squareerror = 0
            for T, Ndop, doptype, nxc, ve, vh, ni, n, p, taum, pp in zip(Tlist, Doplist,Typelist, nxclist, velist, vhlist, nilist, nlist, plist, taulist,[1,fraction*100]):
                n1 = ni * np.exp(Et1 / kb / T)
                p1 = ni * np.exp(-Et1 / kb / T)
                n2 = ni * np.exp(Et2 / kb / T)
                p2 = ni * np.exp(-Et2 / kb / T)
                R = (n * p - ni**2) *((ve*vh/taun1/ve300/(k1*ve*n+vh*p1))+(ve*vh/taun2/ve300/(k2*ve*n2+vh*p)))/(1+((k1*ve*n1+vh*p)/(k1*ve*n+vh*p1))+((k2*ve*n+vh*p2)/(k2*ve*n2+vh*p)))
                taus = pp* nxc / R
                squareerror += np.sum((taus-taum)*(taus-taum)/taus/taus)/len(nxc)
            return squareerror
        
        res0=100
        for i in range(100):    
            try:
                fitres = minimize(squareresidual, x0=[10**(np.random.sample()*10-5),10**(np.random.sample()*10-5),10**(np.random.sample()*10-5),10**(np.random.sample()*10-5),10**(np.random.sample()*10-5)],bounds=[(1e-5,1e5),(1e-5,1e5),(1e-5,1e5),(1e-5,1e5),(1e-5,1e5)],method='L-BFGS-B', options={'eps':1e-9,'gtol':1e-12})
                if fitres.fun < res0:
                    res0 = fitres.fun
                    truefitres = fitres
            except RuntimeError:
                pass
        try:
            taun1list.append(truefitres.x[0]*1e-5)
            taun2list.append(truefitres.x[1]*1e-5)
            k1list.append(truefitres.x[2])
            k2list.append(truefitres.x[3])
            scalelist.append(truefitres.x[4]*100)
            residuallist.append(truefitres.fun)
            del truefitres
        except NameError:
            taun1list.append(0)
            taun2list.append(0)
            k1list.append(0)
            k2list.append(0)
            scalelist.append(1)
            residuallist.append(100)
        count += 1
        print(str(count)+'/'+str(len(Et1list)*len(Et2list))+'finished')

taun1list=np.asarray(taun1list)
taun2list=np.asarray(taun2list)
k1list=np.asarray(k1list)
k2list=np.asarray(k2list)
scalelist=np.asarray(scalelist)
residuallist=np.asarray(residuallist)

optind = np.argmin(residuallist)
Et1opt= Etlist[optind][0]
Et2opt=Etlist[optind][1]
taun1opt= taun1list[optind]
taun2opt= taun2list[optind]
k1opt= k1list[optind]
k2opt= k2list[optind]
scaleopt= scalelist[optind]
resiopt=residuallist[optind]

#Et2opt = -0.3
#Et1opt = 0.15
#se1 = 1e-14
#k1opt = 18
#se2 = se1/103.2
#k2opt = 86
#Ntlist = np.asarray([1e14,1e13])
#resiopt=residuallist[optind]


plt.figure('fitting')

for taum, T, Ndop, doptype, nxc, p in zip(taulist, Tlist,Doplist,Typelist, nxclist, [1,scaleopt]):

    lm=plt.plot(nxc,taum,'.')
    tauopt=p*twolevelSRH_full(nxc=nxc,Ndop=Ndop, T=T, doptype=doptype, Nt=Nt, Et1=Et1opt, Et2=Et2opt,sigma_e1=1/taun1opt/Nt/ve300, sigma_h1=1/taun1opt/Nt/ve300/k1opt,sigma_e2=1/taun2opt/Nt/ve300, sigma_h2=1/taun2opt/Nt/ve300/k2opt)
    plt.plot(nxc,tauopt,'--',color=lm[0].get_color())

plt.title('Residual={:.5e}'.format(resiopt))
plt.loglog()
plt.xlabel(r'Excess carrier density [$\rm cm^{-3}$]')
plt.ylabel(r'Lifetime [s]')

taun1listr = np.reshape(taun1list,(len(Et1list),len(Et2list)))
taun2listr = np.reshape(taun2list,(len(Et1list),len(Et2list)))
k1listr = np.reshape(k1list,(len(Et1list),len(Et2list)))
k2listr = np.reshape(k2list,(len(Et1list),len(Et2list)))
residuallistr = np.reshape(residuallist,(len(Et1list),len(Et2list)))
scalelistr = np.reshape(scalelist,(len(Et1list),len(Et2list)))

extent = (Et2list[0]-(Et2list[1]-Et2list[0])/2, Et2list[-1]+(Et2list[-1]-Et2list[-2])/2, Et1list[0]-(Et1list[1]-Et1list[0])/2, Et1list[-1]+(Et1list[-1]-Et1list[-2])/2)

plt.figure('Resudual')
im1 = plt.imshow(residuallistr,extent =extent, aspect='equal', origin = 'lower',norm=colors.LogNorm())
plt.colorbar(im1)
plt.plot(Etlist[optind][1],Etlist[optind][0],'bo')
plt.plot([-0.3],[0.15],'r*')
plt.xlabel(r'$E_{t2}-E_{i} \/\/ \rm [eV]$')
plt.ylabel(r'$E_{t1}-E_{i} \/\/ \rm [eV]$')

np.save('taun1list.npy',taun1list)
np.save('taun2list.npy',taun2list)
np.save('taun1listr.npy',taun1listr)
np.save('taun2listr.npy',taun2listr)
np.save('k1list.npy',k1list)
np.save('k2list.npy',k2list)
np.save('k1listr.npy',k1listr)
np.save('k2listr.npy',k2listr)
np.save('residuallist.npy',residuallist)
np.save('residuallistr.npy',residuallistr)
np.save('scalelist.npy',scalelist)
np.save('scalelistr.npy',scalelistr)
np.save('Etlist.npy',Etlist)


plt.show()



